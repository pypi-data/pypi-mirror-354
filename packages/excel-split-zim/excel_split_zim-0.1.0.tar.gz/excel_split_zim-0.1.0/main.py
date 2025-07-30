import os
from openpyxl import load_workbook

def split_excel_bysheet(file_path):
    # 获取输入文件所在目录
    output_dir = os.path.dirname(file_path)
    
    # 加载Excel文件
    workbook = load_workbook(file_path)
    
    # 遍历该文件中的所有工作表
    for sheet_name in workbook.sheetnames:
        # 创建新的工作簿
        new_workbook = load_workbook(file_path)
        
        # 删除不需要的工作表
        for sheet in new_workbook.sheetnames:
            if sheet != sheet_name:
                del new_workbook[sheet]
        
        # 保存新工作簿
        output_path = os.path.join(output_dir, f"{sheet_name}.xlsx")
        new_workbook.save(output_path)
        print(f"已保存工作表: {sheet_name}")

file_path = "C:/Users/intern-2/Desktop/案例/达观数据-爱康体检落单0422 2.xlsx"
split_excel_bysheet(file_path)