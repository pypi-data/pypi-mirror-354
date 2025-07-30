import logging
import json
import hashlib
import asyncio
import aiohttp
from typing import Dict, List, Optional, Any, Callable
from pathlib import Path
import os
import time
import requests
from bs4 import BeautifulSoup
from urllib.parse import urljoin, urlparse
from urllib.robotparser import RobotFileParser
from dataclasses import dataclass, field
from collections import deque, defaultdict, Counter
import re
from datetime import datetime

from .analyzers.meta_analyzer import MetaAnalyzer
from .analyzers.content_analyzer import ContentAnalyzer
from .analyzers.modern_seo_analyzer import ModernSEOAnalyzer
from .analyzers.competitive_analyzer import CompetitiveAnalyzer
from .analyzers.advanced_analyzer import AdvancedSEOAnalyzer
from .reporting.report_formatter import ReportFormatter
from .reporting.detailed_report import DetailedReport
from .utils.cache_manager import CacheManager
from .utils.error_handler import setup_logging, handle_analysis_error, TFQ0SEOError
from .utils.paths import TFQSEO_HOME

# Setup logging
setup_logging(TFQSEO_HOME / 'tfq0seo.log')

@dataclass
class CrawlConfig:
    """Configuration for site crawling."""
    max_depth: int = 3
    max_pages: int = 500
    concurrent_requests: int = 10
    delay_between_requests: float = 0.5
    follow_redirects: bool = True
    respect_robots_txt: bool = True
    include_external_links: bool = True
    crawl_images: bool = True
    crawl_css: bool = False
    crawl_js: bool = False
    user_agent: str = "TFQ0SEO-Spider/2.0 (+https://github.com/tfq0/tfq0seo)"
    timeout: int = 30
    max_retries: int = 3
    allowed_domains: List[str] = field(default_factory=list)
    excluded_paths: List[str] = field(default_factory=list)
    include_query_params: bool = False
    custom_headers: Dict[str, str] = field(default_factory=dict)

@dataclass
class CrawlResult:
    """Data structure for crawl results."""
    url: str
    status_code: int
    redirect_url: Optional[str] = None
    redirect_chain: List[str] = field(default_factory=list)
    response_time: float = 0.0
    content_type: str = ""
    content_length: int = 0
    title: str = ""
    meta_description: str = ""
    h1_tags: List[str] = field(default_factory=list)
    h2_tags: List[str] = field(default_factory=list)
    canonical_url: Optional[str] = None
    robots_meta: str = ""
    internal_links: List[str] = field(default_factory=list)
    external_links: List[str] = field(default_factory=list)
    images: List[Dict] = field(default_factory=list)
    schema_markup: List[Dict] = field(default_factory=list)
    word_count: int = 0
    errors: List[str] = field(default_factory=list)
    crawl_depth: int = 0
    parent_url: Optional[str] = None

# Default settings that were previously in config
DEFAULT_SETTINGS = {
    'version': '2.0.0',  # Track settings version with enhanced features
    'seo_thresholds': {
        'title_length': {'min': 30, 'max': 60},
        'meta_description_length': {'min': 120, 'max': 160},
        'content_length': {'min': 300},
        'sentence_length': {'max': 20},
        'keyword_density': {'max': 3.0}
    },
    'readability_thresholds': {
        'flesch_reading_ease': {'min': 60},
        'gunning_fog': {'max': 12}
    },
    'crawling': {
        'timeout': 30,
        'max_retries': 3,
        'user_agent': 'TFQ0SEO-Spider/2.0',
        'max_depth': 3,
        'max_pages': 500,
        'concurrent_requests': 10,
        'delay_between_requests': 0.5,
        'respect_robots_txt': True
    },
    'cache': {
        'enabled': True,
        'expiration': 3600,
        'directory': str(TFQSEO_HOME / 'cache')  # Absolute path in user's home
    },
    'logging': {
        'level': 'INFO',
        'file': str(TFQSEO_HOME / 'tfq0seo.log'),  # Absolute path in user's home
        'format': '%(asctime)s - %(name)s - %(levelname)s - %(message)s',
        'max_size': 10485760,  # 10MB
        'backup_count': 5
    }
}

class SEOAnalyzerApp:
    """tfq0seo main application class.
    
    Provides comprehensive SEO analysis capabilities:
    - URL analysis
    - Content optimization
    - Meta tag validation
    - Modern SEO features
    - Competitive analysis
    - Advanced SEO features
    - Educational resources
    
    Features:
    - Caching support
    - Multiple export formats
    - Detailed recommendations
    - Educational content
    """
    
    def __init__(self):
        """Initialize the tfq0seo application."""
        self.settings = DEFAULT_SETTINGS
        
        # Set up logging
        log_file = TFQSEO_HOME / 'tfq0seo.log'
        setup_logging(log_file)
        self.logger = logging.getLogger('tfq0seo')
        
        # Initialize analyzers
        self.meta_analyzer = MetaAnalyzer(self.settings)
        self.content_analyzer = ContentAnalyzer(self.settings)
        self.modern_analyzer = ModernSEOAnalyzer(self.settings)
        self.competitive_analyzer = CompetitiveAnalyzer(self.settings)
        self.advanced_analyzer = AdvancedSEOAnalyzer(self.settings)
        
        # Create analyzers dictionary for easy access
        self.analyzers = {
            'basic_seo': self.meta_analyzer,
            'content': self.content_analyzer,
            'modern_seo': self.modern_analyzer,
            'competitive': self.competitive_analyzer,
            'advanced_seo': self.advanced_analyzer
        }
        
        # Initialize cache
        self.cache = CacheManager(
            self.settings['cache']['directory'], 
            self.settings['cache']['expiration']
        )
        
        # Crawl state for site crawling
        self.crawled_urls: set = set()
        self.failed_urls: set = set()
        self.redirected_urls: Dict[str, str] = {}
        self.robots_parsers: Dict[str, RobotFileParser] = {}
        self.crawl_results: Dict[str, CrawlResult] = {}
        self.sitemap_urls: set = set()
        
        # Statistics
        self.crawl_stats = {
            'total_urls_found': 0,
            'total_urls_crawled': 0,
            'total_errors': 0,
            'start_time': 0,
            'end_time': 0,
            'status_codes': defaultdict(int),
            'content_types': defaultdict(int)
        }

    def analyze_url(self, url: str) -> Dict[str, Any]:
        """Analyze a single URL for SEO optimization opportunities.
        
        Performs comprehensive real-time SEO analysis including:
        - Technical SEO elements
        - Content quality and optimization
        - Modern SEO features
        - Competitive analysis
        - Advanced SEO techniques
        
        Args:
            url: The URL to analyze
            
        Returns:
            Dict containing comprehensive analysis results
        """
        # Import requests locally to avoid namespace issues
        import requests
        
        try:
            # Fetch and parse the webpage
            response = requests.get(url, headers={'User-Agent': self.settings['crawling']['user_agent']})
            response.raise_for_status()
            
            soup = BeautifulSoup(response.text, 'html.parser')
            
            # Extract meaningful content for analysis
            meaningful_content = self._extract_meaningful_content(soup)
            
            # Perform real-time analysis using all analyzers
            analysis_modules = {}
            
            # Basic SEO Analysis
            basic_seo = self.analyzers['basic_seo'].analyze(soup)
            analysis_modules['basic_seo'] = basic_seo
            
            # Content Analysis with meaningful content
            if meaningful_content:
                content_analysis = self.analyzers['content'].analyze(meaningful_content)
                analysis_modules['content'] = content_analysis
            else:
                analysis_modules['content'] = {
                    'basic_metrics': {'word_count': 0},
                    'error': 'No meaningful content found'
                }
            
            # Modern SEO Analysis
            modern_seo = self.analyzers['modern_seo'].analyze(url)
            analysis_modules['modern_seo'] = modern_seo
            
            # Advanced SEO Analysis
            advanced_seo = self.analyzers['advanced_seo'].analyze(url)
            analysis_modules['advanced_seo'] = advanced_seo
            
            # Competitive Analysis
            competitive = self.analyzers['competitive'].analyze(url)
            analysis_modules['competitive'] = competitive
            
            # Calculate dynamic SEO score
            seo_score = self._calculate_seo_score({'analysis_modules': analysis_modules})
            
            # Generate real-time insights
            insights = self._generate_real_time_insights(analysis_modules)
            
            # Combine all analysis results in the expected format
            combined_report = {
                'url': url,
                'analysis_modules': analysis_modules,
                'timestamp': datetime.now().isoformat(),
                'tool_version': self.settings['version'],
                'seo_score': seo_score,
                'insights': insights,
                'scores': {
                    'overall_score': seo_score,
                    'category_scores': {
                        'technical_seo': min(35, self._calculate_technical_score(analysis_modules)),
                        'content_quality': min(35, self._calculate_content_score(analysis_modules)),
                        'user_experience': min(20, self._calculate_ux_score(analysis_modules)),
                        'security': min(10, self._calculate_security_score(analysis_modules))
                    }
                },
                'summary': {
                    'overview': {
                        'total_issues': len(insights.get('critical_issues', [])) + len(insights.get('opportunities', [])),
                        'critical_issues': len(insights.get('critical_issues', [])),
                        'strongest_category': self._get_strongest_category(analysis_modules),
                        'weakest_category': self._get_weakest_category(analysis_modules)
                    },
                    'key_findings': insights.get('critical_issues', [])[:3] + insights.get('opportunities', [])[:2]
                }
            }
            
            return combined_report
            
        except requests.RequestException as e:
            return {
                'error': f'Failed to fetch URL: {str(e)}',
                'url': url,
                'timestamp': datetime.now().isoformat()
            }
        except Exception as e:
            return {
                'error': f'Analysis failed: {str(e)}',
                'url': url,
                'timestamp': datetime.now().isoformat()
            }

    def _generate_real_time_insights(self, analysis_modules: Dict) -> Dict:
        """Generate real-time insights based on actual analysis results."""
        insights = {
            'critical_issues': [],
            'opportunities': [],
            'strengths': [],
            'recommendations': []
        }
        
        # Analyze HTTPS/Security
        security = analysis_modules.get('modern_seo', {}).get('security', {})
        if not security.get('https'):
            insights['critical_issues'].append("Site is not using HTTPS - major security and SEO issue")
            insights['recommendations'].append("Migrate to HTTPS immediately to improve security and search rankings")
        elif not security.get('ssl_certificate_valid'):
            insights['critical_issues'].append("SSL certificate is invalid or expired")
        else:
            insights['strengths'].append("Site properly uses HTTPS with valid SSL certificate")
        
        # Analyze Content Quality
        content = analysis_modules.get('content', {})
        word_count = content.get('basic_metrics', {}).get('word_count', 0)
        
        if word_count == 0:
            insights['critical_issues'].append("No meaningful content detected - page may be JavaScript-heavy or have extraction issues")
        elif word_count < 300:
            insights['opportunities'].append(f"Content is thin ({word_count} words) - consider expanding to 300+ words")
        elif word_count >= 1000:
            insights['strengths'].append(f"Good content length ({word_count} words)")
        
        # Check for keyword stuffing
        keyword_stuffing = content.get('keyword_analysis', {}).get('keyword_stuffing_detected', [])
        if keyword_stuffing:
            insights['critical_issues'].append(f"Keyword stuffing detected: {', '.join([kw['keyword'] for kw in keyword_stuffing[:3]])}")
            insights['recommendations'].append("Reduce keyword density and focus on natural language")
        
        # Analyze Technical SEO
        basic_seo = analysis_modules.get('basic_seo', {})
        
        # Title analysis
        title = basic_seo.get('title', '')
        if not title:
            insights['critical_issues'].append("Missing title tag")
        elif len(title) > 60:
            insights['opportunities'].append(f"Title tag too long ({len(title)} chars) - optimize to under 60 characters")
        elif 30 <= len(title) <= 60:
            insights['strengths'].append("Title tag length is optimal")
        
        # Meta description analysis
        meta_desc = basic_seo.get('meta_description', '')
        if not meta_desc:
            insights['opportunities'].append("Missing meta description - add one to improve click-through rates")
        elif len(meta_desc) > 160:
            insights['opportunities'].append(f"Meta description too long ({len(meta_desc)} chars)")
        elif 120 <= len(meta_desc) <= 160:
            insights['strengths'].append("Meta description length is optimal")
        
        # H1 analysis
        h1_tags = basic_seo.get('headings', {}).get('h1', [])
        if not h1_tags:
            insights['opportunities'].append("Missing H1 tag - add one for better content structure")
        elif len(h1_tags) > 1:
            insights['opportunities'].append(f"Multiple H1 tags ({len(h1_tags)}) - use only one H1 per page")
        else:
            insights['strengths'].append("Proper H1 tag structure")
        
        # Mobile optimization
        mobile = analysis_modules.get('modern_seo', {}).get('mobile_friendly', {})
        if not mobile.get('viewport_meta'):
            insights['critical_issues'].append("Missing viewport meta tag - critical for mobile optimization")
        elif mobile.get('responsive_viewport'):
            insights['strengths'].append("Mobile-responsive viewport configured")
        
        # Schema markup
        schema_count = len(analysis_modules.get('modern_seo', {}).get('structured_data', {}).get('schema_types', []))
        if schema_count == 0:
            insights['opportunities'].append("No structured data found - implement Schema.org markup for better search visibility")
        else:
            insights['strengths'].append(f"Structured data implemented ({schema_count} schema types)")
        
        return insights

    @handle_analysis_error
    async def crawl_and_analyze_site(self, start_url: str, crawl_config: CrawlConfig = None, 
                                   progress_callback: Callable = None) -> Dict[str, Any]:
        """
        Perform comprehensive site crawl and analysis.
        
        Args:
            start_url: Starting URL for the crawl
            crawl_config: Crawl configuration options
            progress_callback: Callback for progress updates
            
        Returns:
            Comprehensive analysis results
        """
        self.logger.info(f"Starting comprehensive site crawl for: {start_url}")
        
        # Use default config if none provided
        if crawl_config is None:
            crawl_config = CrawlConfig(
                max_depth=self.settings['crawling']['max_depth'],
                max_pages=self.settings['crawling']['max_pages'],
                concurrent_requests=self.settings['crawling']['concurrent_requests'],
                respect_robots_txt=self.settings['crawling']['respect_robots_txt']
            )
        
        # Clear previous crawl state
        self._reset_crawl_state()
        
        # Perform site crawl
        crawl_results = await self._crawl_site(start_url, crawl_config, progress_callback)
        
        # Analyze crawled data
        analysis_results = await self._analyze_crawl_results(crawl_results, progress_callback)
        
        # Generate comprehensive report
        comprehensive_report = self._generate_comprehensive_report(crawl_results, analysis_results)
        
        return comprehensive_report

    async def _crawl_site(self, start_url: str, config: CrawlConfig, 
                         progress_callback: Callable = None) -> Dict[str, Any]:
        """Crawl an entire website starting from the given URL."""
        self.crawl_stats['start_time'] = time.time()
        
        # Parse base domain
        parsed_url = urlparse(start_url)
        base_domain = f"{parsed_url.scheme}://{parsed_url.netloc}"
        
        # Add base domain to allowed domains if not specified
        if not config.allowed_domains:
            config.allowed_domains.append(parsed_url.netloc)
        
        # Load robots.txt if respecting it
        if config.respect_robots_txt:
            await self._load_robots_txt(base_domain)
        
        # Load sitemaps
        await self._load_sitemaps(base_domain)
        
        # Initialize crawl queue
        crawl_queue = deque([(start_url, 0, None)])  # (url, depth, parent_url)
        
        # Start crawling
        async with aiohttp.ClientSession(
            timeout=aiohttp.ClientTimeout(total=config.timeout),
            connector=aiohttp.TCPConnector(limit=config.concurrent_requests),
            headers={'User-Agent': config.user_agent}
        ) as session:
            
            # Process crawl queue
            while crawl_queue and len(self.crawled_urls) < config.max_pages:
                
                # Get batch of URLs to process
                batch = []
                for _ in range(min(config.concurrent_requests, len(crawl_queue))):
                    if crawl_queue:
                        batch.append(crawl_queue.popleft())
                
                if not batch:
                    break
                
                # Process batch concurrently
                tasks = [
                    self._crawl_url(session, url, depth, parent_url, config) 
                    for url, depth, parent_url in batch
                ]
                
                results = await asyncio.gather(*tasks, return_exceptions=True)
                
                # Process results and add new URLs to queue
                for i, result in enumerate(results):
                    if isinstance(result, Exception):
                        url = batch[i][0]
                        self.logger.error(f"Error crawling {url}: {str(result)}")
                        self.failed_urls.add(url)
                        continue
                    
                    if result:
                        url, depth, parent_url = batch[i]
                        
                        # Add newly discovered URLs to queue
                        if depth < config.max_depth:
                            new_urls = result.internal_links
                            for new_url in new_urls:
                                if (new_url not in self.crawled_urls and 
                                    new_url not in self.failed_urls and
                                    self._should_crawl_url(new_url, config)):
                                    crawl_queue.append((new_url, depth + 1, url))
                        
                        # Update progress
                        if progress_callback:
                            progress_callback({
                                'stage': 'crawling',
                                'crawled': len(self.crawled_urls),
                                'queue_size': len(crawl_queue),
                                'current_url': url
                            })
                
                # Respect delay between requests
                if config.delay_between_requests > 0:
                    await asyncio.sleep(config.delay_between_requests)
        
        self.crawl_stats['end_time'] = time.time()
        self.crawl_stats['total_urls_crawled'] = len(self.crawled_urls)
        self.crawl_stats['total_errors'] = len(self.failed_urls)
        
        return self._compile_crawl_results()

    async def _crawl_url(self, session: aiohttp.ClientSession, url: str, depth: int, 
                        parent_url: str, config: CrawlConfig) -> Optional[CrawlResult]:
        """Crawl a single URL and extract SEO data."""
        if url in self.crawled_urls:
            return None
        
        # Check robots.txt
        if not self._is_allowed_by_robots(url, config):
            return None
        
        start_time = time.time()
        
        try:
            # Add custom headers
            headers = dict(config.custom_headers)
            
            async with session.get(url, headers=headers) as response:
                response_time = time.time() - start_time
                content = await response.text()
                
                # Create crawl result
                result = CrawlResult(
                    url=url,
                    status_code=response.status,
                    response_time=response_time,
                    content_type=response.headers.get('content-type', ''),
                    content_length=len(content),
                    crawl_depth=depth,
                    parent_url=parent_url
                )
                
                # Handle redirects
                if response.status in [301, 302, 303, 307, 308]:
                    result.redirect_url = str(response.url)
                    result.redirect_chain = [url, str(response.url)]
                
                # Parse HTML content for SEO data
                if 'text/html' in result.content_type.lower():
                    soup = BeautifulSoup(content, 'html.parser')
                    await self._extract_seo_data(result, soup, url)
                
                # Update statistics
                self.crawl_stats['status_codes'][response.status] += 1
                self.crawl_stats['content_types'][result.content_type] += 1
                
                # Store result
                self.crawled_urls.add(url)
                self.crawl_results[url] = result
                
                return result
                
        except Exception as e:
            self.logger.error(f"Error crawling {url}: {str(e)}")
            self.failed_urls.add(url)
            return None

    async def _extract_seo_data(self, result: CrawlResult, soup: BeautifulSoup, base_url: str):
        """Extract comprehensive SEO data from HTML."""
        
        # Title
        title_tag = soup.find('title')
        result.title = title_tag.get_text().strip() if title_tag else ""
        
        # Meta description
        meta_desc = soup.find('meta', attrs={'name': 'description'})
        result.meta_description = meta_desc.get('content', '') if meta_desc else ""
        
        # Headings
        result.h1_tags = [h1.get_text().strip() for h1 in soup.find_all('h1')]
        result.h2_tags = [h2.get_text().strip() for h2 in soup.find_all('h2')]
        
        # Canonical URL
        canonical = soup.find('link', attrs={'rel': 'canonical'})
        result.canonical_url = canonical.get('href') if canonical else None
        
        # Robots meta
        robots_meta = soup.find('meta', attrs={'name': 'robots'})
        result.robots_meta = robots_meta.get('content', '') if robots_meta else ""
        
        # Links analysis
        await self._analyze_page_links(result, soup, base_url)
        
        # Images analysis
        await self._analyze_page_images(result, soup, base_url)
        
        # Schema markup
        await self._analyze_schema_markup(result, soup)
        
        # Word count - Extract meaningful content only
        text_content = self._extract_meaningful_content(soup)
        result.word_count = self._count_meaningful_words(text_content)

    async def _analyze_page_links(self, result: CrawlResult, soup: BeautifulSoup, base_url: str):
        """Analyze internal and external links."""
        parsed_base = urlparse(base_url)
        base_domain = parsed_base.netloc
        
        for link in soup.find_all('a', href=True):
            href = link.get('href')
            if not href:
                continue
            
            # Resolve relative URLs
            absolute_url = urljoin(base_url, href)
            parsed_link = urlparse(absolute_url)
            
            # Skip non-HTTP links
            if parsed_link.scheme not in ['http', 'https']:
                continue
            
            # Categorize as internal or external
            if parsed_link.netloc == base_domain:
                if absolute_url not in result.internal_links:
                    result.internal_links.append(absolute_url)
            else:
                result.external_links.append(absolute_url)

    async def _analyze_page_images(self, result: CrawlResult, soup: BeautifulSoup, base_url: str):
        """Analyze images for SEO optimization."""
        for img in soup.find_all('img'):
            src = img.get('src')
            if not src:
                continue
            
            # Resolve relative URLs
            absolute_url = urljoin(base_url, src)
            
            image_data = {
                'url': absolute_url,
                'alt_text': img.get('alt', ''),
                'title': img.get('title', ''),
                'width': img.get('width'),
                'height': img.get('height'),
                'loading': img.get('loading', ''),
                'has_alt': bool(img.get('alt')),
                'is_decorative': not bool(img.get('alt')) and img.get('role') == 'presentation'
            }
            
            result.images.append(image_data)

    async def _analyze_schema_markup(self, result: CrawlResult, soup: BeautifulSoup):
        """Analyze structured data/schema markup."""
        # JSON-LD
        for script in soup.find_all('script', type='application/ld+json'):
            try:
                schema_data = json.loads(script.string)
                result.schema_markup.append({
                    'type': 'json-ld',
                    'data': schema_data
                })
            except:
                continue
        
        # Microdata
        for element in soup.find_all(attrs={'itemtype': True}):
            result.schema_markup.append({
                'type': 'microdata',
                'itemtype': element.get('itemtype'),
                'element': element.name
            })

    async def _load_robots_txt(self, base_url: str):
        """Load and parse robots.txt."""
        robots_url = f"{base_url}/robots.txt"
        try:
            rp = RobotFileParser()
            rp.set_url(robots_url)
            rp.read()
            self.robots_parsers[base_url] = rp
            self.logger.info(f"Loaded robots.txt from {robots_url}")
        except Exception as e:
            self.logger.warning(f"Could not load robots.txt from {robots_url}: {str(e)}")

    async def _load_sitemaps(self, base_url: str):
        """Load and parse XML sitemaps."""
        sitemap_urls = [
            f"{base_url}/sitemap.xml",
            f"{base_url}/sitemap_index.xml"
        ]
        
        for sitemap_url in sitemap_urls:
            try:
                async with aiohttp.ClientSession() as session:
                    async with session.get(sitemap_url) as response:
                        if response.status == 200:
                            content = await response.text()
                            await self._parse_sitemap(content)
                            self.logger.info(f"Loaded sitemap from {sitemap_url}")
                            break
            except Exception as e:
                self.logger.warning(f"Could not load sitemap from {sitemap_url}: {str(e)}")

    async def _parse_sitemap(self, sitemap_content: str):
        """Parse XML sitemap and extract URLs."""
        try:
            import xml.etree.ElementTree as ET
            root = ET.fromstring(sitemap_content)
            
            # Handle sitemap index
            if 'sitemapindex' in root.tag.lower():
                for sitemap in root.findall('.//{http://www.sitemaps.org/schemas/sitemap/0.9}sitemap'):
                    loc = sitemap.find('{http://www.sitemaps.org/schemas/sitemap/0.9}loc')
                    if loc is not None:
                        await self._load_sitemap_url(loc.text)
            
            # Handle regular sitemap
            else:
                for url in root.findall('.//{http://www.sitemaps.org/schemas/sitemap/0.9}url'):
                    loc = url.find('{http://www.sitemaps.org/schemas/sitemap/0.9}loc')
                    if loc is not None:
                        self.sitemap_urls.add(loc.text)
                        
        except Exception as e:
            self.logger.warning(f"Could not parse sitemap XML: {str(e)}")

    async def _load_sitemap_url(self, sitemap_url: str):
        """Load a specific sitemap URL."""
        try:
            async with aiohttp.ClientSession() as session:
                async with session.get(sitemap_url) as response:
                    if response.status == 200:
                        content = await response.text()
                        await self._parse_sitemap(content)
        except Exception as e:
            self.logger.warning(f"Could not load sitemap from {sitemap_url}: {str(e)}")

    def _should_crawl_url(self, url: str, config: CrawlConfig) -> bool:
        """Determine if a URL should be crawled based on configuration."""
        parsed_url = urlparse(url)
        
        # Check allowed domains
        if config.allowed_domains:
            if parsed_url.netloc not in config.allowed_domains:
                return False
        
        # Check excluded paths
        for excluded_path in config.excluded_paths:
            if excluded_path in parsed_url.path:
                return False
        
        # Check file extensions (skip non-HTML by default)
        path = parsed_url.path.lower()
        if path.endswith(('.pdf', '.doc', '.docx', '.xls', '.xlsx', '.ppt', '.pptx', '.zip', '.rar')):
            return False
        
        return True

    def _is_allowed_by_robots(self, url: str, config: CrawlConfig) -> bool:
        """Check if URL is allowed by robots.txt."""
        if not config.respect_robots_txt:
            return True
        
        parsed_url = urlparse(url)
        base_url = f"{parsed_url.scheme}://{parsed_url.netloc}"
        
        if base_url in self.robots_parsers:
            rp = self.robots_parsers[base_url]
            return rp.can_fetch(config.user_agent, url)
        
        return True

    def _compile_crawl_results(self) -> Dict[str, Any]:
        """Compile comprehensive crawl results."""
        return {
            'summary': {
                'total_urls_crawled': len(self.crawled_urls),
                'total_urls_failed': len(self.failed_urls),
                'crawl_duration': self.crawl_stats['end_time'] - self.crawl_stats['start_time'],
                'sitemap_urls_found': len(self.sitemap_urls)
            },
            'statistics': self.crawl_stats,
            'pages': {url: result.__dict__ for url, result in self.crawl_results.items()},
            'failed_urls': list(self.failed_urls),
            'redirects': self.redirected_urls,
            'sitemap_urls': list(self.sitemap_urls)
        }

    async def _analyze_crawl_results(self, crawl_results: Dict[str, Any], 
                                   progress_callback: Callable = None) -> Dict[str, Any]:
        """Analyze the results from site crawling."""
        analysis_results = {
            'url_analysis': {},
            'content_analysis': {},
            'duplicate_content': [],
            'broken_links': [],
            'redirect_chains': [],
            'sitemap_coverage': {},
            'summary_statistics': {}
        }
        
        pages = crawl_results.get('pages', {})
        total_pages = len(pages)
        
        # Analyze each page
        for i, (url, page_data) in enumerate(pages.items()):
            if progress_callback:
                progress_callback({
                    'stage': 'analyzing',
                    'current': i + 1,
                    'total': total_pages,
                    'url': url
                })
            
            # URL Analysis
            url_analysis = self._analyze_url_structure(url)
            analysis_results['url_analysis'][url] = url_analysis
            
            # Content Analysis (if HTML content available)
            if page_data.get('status_code') == 200 and 'text/html' in page_data.get('content_type', ''):
                content_analysis = self._analyze_page_content_detailed(url, page_data)
                analysis_results['content_analysis'][url] = content_analysis
        
        # Site-wide analysis
        analysis_results['site_wide_analysis'] = self._perform_site_wide_analysis(pages)
        analysis_results['summary_statistics'] = self._generate_summary_statistics(analysis_results)
        
        return analysis_results

    def _analyze_url_structure(self, url: str) -> Dict[str, Any]:
        """Analyze URL structure for SEO optimization."""
        parsed_url = urlparse(url)
        
        # Basic URL analysis
        url_analysis = {
            'url': url,
            'protocol': parsed_url.scheme,
            'domain': parsed_url.netloc,
            'path': parsed_url.path,
            'query': parsed_url.query,
            'fragment': parsed_url.fragment,
            'is_seo_friendly': True,
            'url_length': len(url),
            'path_segments': [seg for seg in parsed_url.path.split('/') if seg],
            'has_query_parameters': bool(parsed_url.query),
            'seo_issues': [],
            'recommendations': []
        }
        
        # URL length check
        if url_analysis['url_length'] > 100:
            url_analysis['seo_issues'].append(f"URL is too long ({url_analysis['url_length']} characters)")
            url_analysis['recommendations'].append("Shorten URL to under 100 characters")
            url_analysis['is_seo_friendly'] = False
        
        # Path depth check
        if len(url_analysis['path_segments']) > 5:
            url_analysis['seo_issues'].append(f"URL has deep path structure ({len(url_analysis['path_segments'])} segments)")
            url_analysis['recommendations'].append("Reduce URL depth for better crawlability")
        
        # Query parameters check
        if url_analysis['has_query_parameters']:
            url_analysis['recommendations'].append("Consider URL rewriting for cleaner URLs")
        
        return url_analysis

    def _analyze_page_content_detailed(self, url: str, page_data: Dict) -> Dict[str, Any]:
        """Analyze content of a single page in detail."""
        try:
            content_analysis = {
                'title': page_data.get('title', ''),
                'meta_description': page_data.get('meta_description', ''),
                'h1_tags': page_data.get('h1_tags', []),
                'h2_tags': page_data.get('h2_tags', []),
                'word_count': page_data.get('word_count', 0),
                'canonical_url': page_data.get('canonical_url'),
                'robots_meta': page_data.get('robots_meta', ''),
                'internal_links_count': len(page_data.get('internal_links', [])),
                'external_links_count': len(page_data.get('external_links', [])),
                'images_count': len(page_data.get('images', [])),
                'schema_markup_count': len(page_data.get('schema_markup', [])),
                'issues': [],
                'recommendations': []
            }
            
            # Content quality checks
            if not content_analysis['title']:
                content_analysis['issues'].append("Missing title tag")
                content_analysis['recommendations'].append("Add a descriptive title tag")
            elif len(content_analysis['title']) > 60:
                content_analysis['issues'].append("Title tag too long")
                content_analysis['recommendations'].append("Shorten title to under 60 characters")
            
            if not content_analysis['meta_description']:
                content_analysis['issues'].append("Missing meta description")
                content_analysis['recommendations'].append("Add a meta description")
            elif len(content_analysis['meta_description']) > 160:
                content_analysis['issues'].append("Meta description too long")
                content_analysis['recommendations'].append("Shorten meta description to under 160 characters")
            
            if len(content_analysis['h1_tags']) == 0:
                content_analysis['issues'].append("Missing H1 tag")
                content_analysis['recommendations'].append("Add an H1 tag for better structure")
            elif len(content_analysis['h1_tags']) > 1:
                content_analysis['issues'].append("Multiple H1 tags found")
                content_analysis['recommendations'].append("Use only one H1 tag per page")
            
            if content_analysis['word_count'] < 300:
                content_analysis['issues'].append("Thin content (less than 300 words)")
                content_analysis['recommendations'].append("Add more content to improve page value")
            
            return content_analysis
            
        except Exception as e:
            self.logger.error(f"Error analyzing content for {url}: {str(e)}")
            return {'error': str(e)}

    def _perform_site_wide_analysis(self, pages: Dict[str, Any]) -> Dict[str, Any]:
        """Perform site-wide analysis across all pages."""
        site_analysis = {
            'duplicate_content': self._find_duplicate_content(pages),
            'broken_links': self._find_broken_links(pages),
            'redirect_chains': self._analyze_redirect_chains(pages),
            'orphaned_pages': self._find_orphaned_pages(pages),
            'site_structure': self._analyze_site_structure(pages),
            'content_gaps': self._analyze_content_gaps(pages),
            'internal_linking': self._analyze_internal_linking(pages)
        }
        
        return site_analysis

    def _find_duplicate_content(self, pages: Dict[str, Any]) -> List[List[str]]:
        """Find potential duplicate content across pages."""
        content_fingerprints = {}
        
        for url, page_data in pages.items():
            if page_data.get('status_code') == 200:
                # Create content fingerprint
                title = page_data.get('title', '')
                meta_desc = page_data.get('meta_description', '')
                h1_tags = ' '.join(page_data.get('h1_tags', []))
                
                fingerprint = f"{title}|{meta_desc}|{h1_tags}".strip()
                
                if fingerprint and fingerprint != "||":
                    if fingerprint in content_fingerprints:
                        content_fingerprints[fingerprint].append(url)
                    else:
                        content_fingerprints[fingerprint] = [url]
        
        # Return groups with duplicates
        return [urls for urls in content_fingerprints.values() if len(urls) > 1]

    def _find_broken_links(self, pages: Dict[str, Any]) -> List[Dict[str, Any]]:
        """Find broken links across the site."""
        broken_links = []
        
        for url, page_data in pages.items():
            if page_data.get('status_code', 0) >= 400:
                broken_links.append({
                    'url': url,
                    'status_code': page_data.get('status_code'),
                    'parent_url': page_data.get('parent_url'),
                    'error_type': self._get_error_type(page_data.get('status_code', 0))
                })
        
        return broken_links

    def _analyze_redirect_chains(self, pages: Dict[str, Any]) -> List[Dict[str, Any]]:
        """Analyze redirect chains."""
        redirect_chains = []
        
        for url, page_data in pages.items():
            redirect_chain = page_data.get('redirect_chain', [])
            if len(redirect_chain) > 1:
                redirect_chains.append({
                    'original_url': redirect_chain[0],
                    'final_url': redirect_chain[-1],
                    'chain_length': len(redirect_chain),
                    'full_chain': redirect_chain,
                    'status_code': page_data.get('status_code')
                })
        
        return redirect_chains

    def _find_orphaned_pages(self, pages: Dict[str, Any]) -> List[str]:
        """Find pages that have no internal links pointing to them."""
        all_urls = set(pages.keys())
        linked_urls = set()
        
        for url, page_data in pages.items():
            internal_links = page_data.get('internal_links', [])
            linked_urls.update(internal_links)
        
        orphaned_pages = all_urls - linked_urls
        
        # Remove the start URL as it's not expected to be linked internally
        start_urls = [url for url in all_urls if pages[url].get('crawl_depth', 1) == 0]
        for start_url in start_urls:
            orphaned_pages.discard(start_url)
        
        return list(orphaned_pages)

    def _analyze_site_structure(self, pages: Dict[str, Any]) -> Dict[str, Any]:
        """Analyze overall site structure."""
        depths = [page_data.get('crawl_depth', 0) for page_data in pages.values()]
        
        return {
            'max_depth': max(depths) if depths else 0,
            'average_depth': sum(depths) / len(depths) if depths else 0,
            'total_pages': len(pages),
            'pages_by_depth': {depth: depths.count(depth) for depth in set(depths)}
        }

    def _analyze_content_gaps(self, pages: Dict[str, Any]) -> Dict[str, Any]:
        """Analyze content gaps and opportunities."""
        word_counts = []
        title_lengths = []
        meta_desc_lengths = []
        
        for url, page_data in pages.items():
            if page_data.get('status_code') == 200:
                word_counts.append(page_data.get('word_count', 0))
                title_lengths.append(len(page_data.get('title', '')))
                meta_desc_lengths.append(len(page_data.get('meta_description', '')))
        
        return {
            'average_word_count': sum(word_counts) / len(word_counts) if word_counts else 0,
            'thin_content_pages': len([wc for wc in word_counts if wc < 300]),
            'missing_titles': len([tl for tl in title_lengths if tl == 0]),
            'missing_meta_descriptions': len([mdl for mdl in meta_desc_lengths if mdl == 0]),
            'title_length_issues': len([tl for tl in title_lengths if tl > 60 or tl < 30]),
            'meta_desc_length_issues': len([mdl for mdl in meta_desc_lengths if mdl > 160 or mdl < 120])
        }

    def _analyze_internal_linking(self, pages: Dict[str, Any]) -> Dict[str, Any]:
        """Analyze internal linking patterns."""
        total_internal_links = 0
        pages_with_internal_links = 0
        
        for url, page_data in pages.items():
            internal_links = page_data.get('internal_links', [])
            total_internal_links += len(internal_links)
            if internal_links:
                pages_with_internal_links += 1
        
        total_pages = len([p for p in pages.values() if p.get('status_code') == 200])
        
        return {
            'total_internal_links': total_internal_links,
            'average_internal_links_per_page': total_internal_links / total_pages if total_pages else 0,
            'pages_with_internal_links': pages_with_internal_links,
            'internal_linking_percentage': (pages_with_internal_links / total_pages * 100) if total_pages else 0
        }

    def _generate_summary_statistics(self, analysis_results: Dict[str, Any]) -> Dict[str, Any]:
        """Generate summary statistics across all analyses."""
        return {
            'total_pages_analyzed': len(analysis_results.get('url_analysis', {})),
            'duplicate_content_groups': len(analysis_results.get('site_wide_analysis', {}).get('duplicate_content', [])),
            'broken_links_found': len(analysis_results.get('site_wide_analysis', {}).get('broken_links', [])),
            'redirect_chains_found': len(analysis_results.get('site_wide_analysis', {}).get('redirect_chains', [])),
            'orphaned_pages': len(analysis_results.get('site_wide_analysis', {}).get('orphaned_pages', []))
        }

    def _generate_comprehensive_report(self, crawl_results: Dict[str, Any], 
                                     analysis_results: Dict[str, Any]) -> Dict[str, Any]:
        """Generate comprehensive analysis report."""
        return {
            'analysis_metadata': {
                'timestamp': datetime.now().isoformat(),
                'analyzer_version': '2.0.0',
                'analysis_type': 'comprehensive_crawl',
                'total_analysis_time': crawl_results.get('summary', {}).get('crawl_duration', 0)
            },
            'crawl_summary': crawl_results.get('summary', {}),
            'crawl_statistics': crawl_results.get('statistics', {}),
            'analysis_results': analysis_results,
            'recommendations': self._generate_prioritized_recommendations(analysis_results),
            'export_data': {
                'pages': crawl_results.get('pages', {}),
                'failed_urls': crawl_results.get('failed_urls', []),
                'sitemap_urls': crawl_results.get('sitemap_urls', [])
            }
        }

    def _generate_prioritized_recommendations(self, analysis_results: Dict[str, Any]) -> List[Dict[str, Any]]:
        """Generate prioritized recommendations based on analysis."""
        recommendations = []
        
        # High priority recommendations
        broken_links = len(analysis_results.get('site_wide_analysis', {}).get('broken_links', []))
        if broken_links > 0:
            recommendations.append({
                'category': 'Critical',
                'priority': 'high',
                'title': f'Fix {broken_links} broken links',
                'description': 'Broken links harm user experience and SEO rankings',
                'impact': 'High',
                'effort': 'Medium'
            })
        
        # Content recommendations
        content_gaps = analysis_results.get('site_wide_analysis', {}).get('content_gaps', {})
        thin_content = content_gaps.get('thin_content_pages', 0)
        if thin_content > 0:
            recommendations.append({
                'category': 'Content',
                'priority': 'high',
                'title': f'Improve {thin_content} pages with thin content',
                'description': 'Pages with less than 300 words may not rank well',
                'impact': 'High',
                'effort': 'High'
            })
        
        missing_meta = content_gaps.get('missing_meta_descriptions', 0)
        if missing_meta > 0:
            recommendations.append({
                'category': 'Meta Data',
                'priority': 'medium',
                'title': f'Add meta descriptions to {missing_meta} pages',
                'description': 'Meta descriptions improve click-through rates from search results',
                'impact': 'Medium',
                'effort': 'Low'
            })
        
        return recommendations

    def _reset_crawl_state(self):
        """Reset crawl state for new crawl."""
        self.crawled_urls.clear()
        self.failed_urls.clear()
        self.redirected_urls.clear()
        self.robots_parsers.clear()
        self.crawl_results.clear()
        self.sitemap_urls.clear()
        
        self.crawl_stats = {
            'total_urls_found': 0,
            'total_urls_crawled': 0,
            'total_errors': 0,
            'start_time': 0,
            'end_time': 0,
            'status_codes': defaultdict(int),
            'content_types': defaultdict(int)
        }

    def _get_error_type(self, status_code: int) -> str:
        """Get human-readable error type for status code."""
        error_types = {
            400: "Bad Request",
            401: "Unauthorized", 
            403: "Forbidden",
            404: "Not Found",
            405: "Method Not Allowed",
            500: "Internal Server Error",
            502: "Bad Gateway",
            503: "Service Unavailable",
            504: "Gateway Timeout"
        }
        return error_types.get(status_code, f"HTTP {status_code}")

    def export_crawl_results(self, format: str = 'json', output_path: str = None) -> str:
        """Export crawl results to various formats."""
        if not self.crawl_results:
            raise TFQ0SEOError(
                error_code="NO_CRAWL_RESULTS",
                message="No crawl results to export"
            )
        
        results = self._compile_crawl_results()
        
        if format == 'json':
            output = json.dumps(results, indent=2, default=str)
        elif format == 'csv':
            output = self._export_to_csv(results)
        elif format == 'xlsx':
            output = self._export_to_xlsx(results, output_path)
            return "XLSX file created"  # Special return for binary files
        else:
            raise ValueError(f"Unsupported export format: {format}")
        
        if output_path:
            with open(output_path, 'w', encoding='utf-8') as f:
                f.write(output)
        
        return output

    def _export_to_csv(self, results: Dict) -> str:
        """Export results to CSV format."""
        import csv
        import io
        
        output = io.StringIO()
        writer = csv.writer(output)
        
        # Header
        header = [
            'URL', 'Status Code', 'Title', 'Meta Description', 'H1 Count', 'H2 Count',
            'Word Count', 'Internal Links', 'External Links', 'Images', 'Response Time',
            'Content Type', 'Canonical URL', 'Robots Meta', 'Depth', 'Parent URL'
        ]
        writer.writerow(header)
        
        # Data rows
        for url, page in results['pages'].items():
            row = [
                page['url'],
                page['status_code'],
                page['title'],
                page['meta_description'],
                len(page['h1_tags']),
                len(page['h2_tags']),
                page['word_count'],
                len(page['internal_links']),
                len(page['external_links']),
                len(page['images']),
                page['response_time'],
                page['content_type'],
                page['canonical_url'] or '',
                page['robots_meta'],
                page['crawl_depth'],
                page['parent_url'] or ''
            ]
            writer.writerow(row)
        
        return output.getvalue()

    def _export_to_xlsx(self, results: Dict, output_path: str) -> str:
        """Export results to XLSX format using openpyxl."""
        try:
            import pandas as pd
            from openpyxl import Workbook
            from openpyxl.styles import Font, PatternFill, Alignment
            from openpyxl.utils.dataframe import dataframe_to_rows
        except ImportError:
            raise ImportError("pandas and openpyxl are required for XLSX export. Install with: pip install pandas openpyxl")
        
        # Create workbook
        wb = Workbook()
        
        # Remove default sheet
        wb.remove(wb.active)
        
        # Create summary sheet
        summary_sheet = wb.create_sheet("Summary")
        summary_data = [
            ["Metric", "Value"],
            ["Total URLs Crawled", results.get('summary', {}).get('total_urls_crawled', 0)],
            ["Total URLs Failed", results.get('summary', {}).get('total_urls_failed', 0)],
            ["Crawl Duration (seconds)", results.get('summary', {}).get('crawl_duration', 0)],
            ["Sitemap URLs Found", results.get('summary', {}).get('sitemap_urls_found', 0)]
        ]
        
        for row in summary_data:
            summary_sheet.append(row)
        
        # Style summary sheet
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        
        for cell in summary_sheet[1]:
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center")
        
        # Create pages sheet
        pages_sheet = wb.create_sheet("Pages")
        
        # Convert pages data to DataFrame
        pages_data = []
        for url, page in results.get('pages', {}).items():
            pages_data.append([
                page.get('url', ''),
                page.get('status_code', ''),
                page.get('title', ''),
                page.get('meta_description', ''),
                len(page.get('h1_tags', [])),
                len(page.get('h2_tags', [])),
                page.get('word_count', 0),
                len(page.get('internal_links', [])),
                len(page.get('external_links', [])),
                len(page.get('images', [])),
                page.get('response_time', 0),
                page.get('content_type', ''),
                page.get('canonical_url', ''),
                page.get('robots_meta', ''),
                page.get('crawl_depth', 0),
                page.get('parent_url', '')
            ])
        
        # Add headers
        headers = [
            'URL', 'Status Code', 'Title', 'Meta Description', 'H1 Count', 'H2 Count',
            'Word Count', 'Internal Links', 'External Links', 'Images', 'Response Time',
            'Content Type', 'Canonical URL', 'Robots Meta', 'Depth', 'Parent URL'
        ]
        
        pages_sheet.append(headers)
        
        # Add data
        for row in pages_data:
            pages_sheet.append(row)
        
        # Style pages sheet headers
        for cell in pages_sheet[1]:
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center")
        
        # Auto-adjust column widths
        for sheet in wb.worksheets:
            for column in sheet.columns:
                max_length = 0
                column_letter = column[0].column_letter
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = min(max_length + 2, 50)  # Cap at 50 characters
                sheet.column_dimensions[column_letter].width = adjusted_width
        
        # Save workbook
        wb.save(output_path)
        return f"XLSX file saved to {output_path}"

    def get_crawl_insights(self) -> Dict[str, Any]:
        """Get quick insights from the crawl results."""
        if not self.crawl_results:
            return {}
        
        results = self._compile_crawl_results()
        broken_links = self._find_broken_links(results['pages'])
        duplicate_content = self._find_duplicate_content(results['pages'])
        
        return {
            'overview': {
                'total_pages': len(self.crawl_results),
                'total_issues': len(broken_links) + len(duplicate_content),
                'crawl_coverage': f"{len(self.crawl_results)} pages crawled"
            },
            'critical_issues': {
                'broken_links': len(broken_links),
                'duplicate_content': len(duplicate_content),
                'failed_urls': len(self.failed_urls)
            },
            'crawl_summary': results['summary']
        }

    @handle_analysis_error
    def analyze_content(self, content: str, target_keyword: Optional[str] = None) -> Dict:
        """Analyze text content for tfq0seo optimization.
        
        Performs content analysis:
        - Keyword optimization
        - Content structure
        - Readability metrics
        - SEO best practices
        
        Args:
            content: Text content to analyze
            target_keyword: Optional focus keyword
            
        Returns:
            Dictionary containing content analysis results
        """
        self.logger.info("Starting content analysis")
        
        # Perform content analysis
        analysis = {
            'content_length': len(content),
            'target_keyword': target_keyword,
            'content_analysis': self.content_analyzer.analyze(content, target_keyword)
        }
        
        return analysis

    def _combine_reports(self, analysis: Dict) -> Dict:
        """Combine analyzer reports into unified tfq0seo report.
        
        Merges results from:
        - Meta analysis
        - Content analysis
        - Modern SEO analysis
        - Competitive analysis (if available)
        - Advanced analysis (if available)
        
        Args:
            analysis: Dictionary containing individual analysis results
            
        Returns:
            Combined report with unified recommendations
        """
        combined = {
            'strengths': [],
            'weaknesses': [],
            'recommendations': [],
            'education_tips': []
        }
        
        # Collect all findings
        for key in ['meta_analysis', 'content_analysis', 'modern_seo_analysis', 'competitive_analysis', 'advanced_analysis']:
            if key in analysis and isinstance(analysis[key], dict):
                report = analysis[key]
                for category in combined.keys():
                    if category in report:
                        combined[category].extend(report[category])
        
        # Remove duplicates while preserving order
        for category in combined.keys():
            combined[category] = list(dict.fromkeys(combined[category]))
        
        # Add summary
        combined['summary'] = {
            'total_strengths': len(combined['strengths']),
            'total_weaknesses': len(combined['weaknesses']),
            'total_recommendations': len(combined['recommendations']),
            'seo_score': self._calculate_seo_score(combined)
        }
        
        return combined

    def _calculate_seo_score(self, report: Dict) -> int:
        """Calculate dynamic SEO score based on actual analysis results."""
        score = 0
        max_score = 100
        
        # Technical SEO (35 points)
        technical_score = 0
        
        # HTTPS check (8 points)
        if report.get('analysis_modules', {}).get('modern_seo', {}).get('security', {}).get('https'):
            technical_score += 8
        
        # Title tag (7 points)
        title = report.get('analysis_modules', {}).get('basic_seo', {}).get('title', '')
        if title and 30 <= len(title) <= 60:
            technical_score += 7
        elif title:
            technical_score += 4
        
        # Meta description (7 points)
        meta_desc = report.get('analysis_modules', {}).get('basic_seo', {}).get('meta_description', '')
        if meta_desc and 120 <= len(meta_desc) <= 160:
            technical_score += 7
        elif meta_desc:
            technical_score += 4
        
        # H1 tags (6 points)
        h1_tags = report.get('analysis_modules', {}).get('basic_seo', {}).get('headings', {}).get('h1', [])
        if len(h1_tags) == 1:
            technical_score += 6
        elif len(h1_tags) > 0:
            technical_score += 3
        
        # URL structure (4 points)
        url_analysis = report.get('analysis_modules', {}).get('advanced_seo', {}).get('url_analysis', {})
        if url_analysis.get('is_seo_friendly'):
            technical_score += 4
        elif url_analysis.get('seo_score', 0) > 50:
            technical_score += 2
        
        # Schema markup (3 points)
        schema_count = len(report.get('analysis_modules', {}).get('modern_seo', {}).get('structured_data', {}).get('schema_types', []))
        if schema_count > 0:
            technical_score += min(3, schema_count)
        
        score += min(35, technical_score)
        
        # Content Quality (35 points)
        content_score = 0
        content_analysis = report.get('analysis_modules', {}).get('content', {})
        
        # Word count (12 points)
        word_count = content_analysis.get('basic_metrics', {}).get('word_count', 0)
        if word_count >= 1000:
            content_score += 12
        elif word_count >= 500:
            content_score += 9
        elif word_count >= 300:
            content_score += 6
        elif word_count > 0:
            content_score += 3
        
        # Readability (10 points)
        flesch_score = content_analysis.get('readability', {}).get('flesch_reading_ease', 0)
        if 60 <= flesch_score <= 80:
            content_score += 10
        elif 50 <= flesch_score <= 90:
            content_score += 8
        elif flesch_score > 0:
            content_score += 4
        
        # Keyword optimization (8 points)
        keyword_analysis = content_analysis.get('keyword_analysis', {})
        if keyword_analysis.get('total_meaningful_words', 0) > 0:
            content_score += 4
            # Check for keyword stuffing (penalty)
            if len(keyword_analysis.get('keyword_stuffing_detected', [])) == 0:
                content_score += 4
            else:
                content_score -= 2  # Penalty for keyword stuffing
        
        # Content structure (5 points)
        structure = content_analysis.get('content_structure', {})
        if structure.get('content_sections', 0) > 2:
            content_score += 3
        if structure.get('avg_paragraph_length', 0) > 0:
            content_score += 2
        
        score += min(35, content_score)
        
        # User Experience (20 points)
        ux_score = 0
        
        # Mobile optimization (12 points)
        mobile_analysis = report.get('analysis_modules', {}).get('modern_seo', {}).get('mobile_friendly', {})
        if mobile_analysis.get('responsive_viewport'):
            ux_score += 6
        if mobile_analysis.get('viewport_meta'):
            ux_score += 4
        if mobile_analysis.get('touch_elements_spacing', {}).get('potentially_small', 0) == 0:
            ux_score += 2
        
        # Performance indicators (8 points)
        html_structure = report.get('analysis_modules', {}).get('modern_seo', {}).get('html_structure', {})
        if html_structure.get('optimizations', {}).get('image_optimization', {}).get('missing_alt', 0) == 0:
            ux_score += 8
        
        score += min(20, ux_score)
        
        # Security (10 points)
        security_score = 0
        security_analysis = report.get('analysis_modules', {}).get('modern_seo', {}).get('security', {})
        
        if security_analysis.get('https') and security_analysis.get('ssl_certificate_valid'):
            security_score += 6
        elif security_analysis.get('https'):
            security_score += 4
        
        if security_analysis.get('hsts'):
            security_score += 2
        if security_analysis.get('xss_protection'):
            security_score += 1
        if security_analysis.get('content_security'):
            security_score += 1
        
        score += min(10, security_score)
        
        # Ensure score is within bounds
        final_score = max(0, min(100, score))
        
        return final_score

    def get_educational_resources(self, topic: Optional[str] = None) -> Dict:
        """Get tfq0seo educational resources.
        
        Provides learning materials on:
        - Meta tags optimization
        - Content SEO strategies
        - Technical SEO implementation
        
        Args:
            topic: Optional specific topic to retrieve
            
        Returns:
            Dictionary containing educational resources
        """
        resources = {
            'meta_tags': [
                {
                    'title': 'Understanding Meta Tags',
                    'description': 'Learn about the importance of meta tags in SEO',
                    'key_points': [
                        'Title tag best practices',
                        'Meta description optimization',
                        'Robots meta directives'
                    ]
                }
            ],
            'content_optimization': [
                {
                    'title': 'Content SEO Guide',
                    'description': 'Best practices for SEO-friendly content',
                    'key_points': [
                        'Keyword research and placement',
                        'Content structure and readability',
                        'Internal linking strategies'
                    ]
                }
            ],
            'technical_seo': [
                {
                    'title': 'Technical SEO Fundamentals',
                    'description': 'Understanding technical aspects of SEO',
                    'key_points': [
                        'Site structure and navigation',
                        'Mobile optimization',
                        'Page speed optimization'
                    ]
                }
            ]
        }
        
        if topic and topic in resources:
            return {topic: resources[topic]}
        return resources

    def get_recommendation_details(self, recommendation: str) -> Dict:
        """Get detailed tfq0seo recommendation information.
        
        Provides:
        - Implementation steps
        - Importance level
        - Resource links
        - Verification process
        
        Args:
            recommendation: The recommendation to get details for
            
        Returns:
            Dictionary containing detailed recommendation information
        """
        return {
            'recommendation': recommendation,
            'importance': 'high',
            'implementation_guide': [
                'Step 1: Understanding the issue',
                'Step 2: Implementation steps',
                'Step 3: Verification process'
            ],
            'resources': [
                'Documentation link 1',
                'Tutorial link 2'
            ]
        }

    def generate_report(self, analysis: Dict, format: str = 'markdown', output_path: Optional[str] = None) -> str:
        """Generate detailed analysis report.
        
        Args:
            analysis: Analysis results dictionary
            format: Output format ('markdown', 'html', 'json', 'csv')
            output_path: Optional path to save report
            
        Returns:
            Formatted report string
        """
        # Generate detailed report
        detailed_report = DetailedReport(analysis)
        report_data = detailed_report.generate_report(format)
        
        # Format report
        formatter = ReportFormatter(report_data)
        return formatter.format_report(format, output_path)

    def export_report(self, analysis: Dict[str, Any], format: str = 'markdown', output_path: Optional[str] = None) -> str:
        """Export analysis report in specified format.
        
        Args:
            analysis: Analysis results to export
            format: Report format ('markdown', 'html', or 'json')
            output_path: Optional path to save report to
        
        Returns:
            Formatted report string
        """
        # Generate detailed report
        detailed_report = DetailedReport(analysis)
        report = detailed_report.generate_report(format)
        
        # Save to file if path provided
        if output_path:
            os.makedirs(os.path.dirname(output_path), exist_ok=True)
            with open(output_path, 'w', encoding='utf-8') as f:
                f.write(report)
        
        return report

    @handle_analysis_error
    def comprehensive_analysis(self, url: str, options: Dict = None) -> Dict:
        """Perform a comprehensive SEO analysis with all available features.
        
        This function combines all analyzers to provide the most detailed analysis:
        - Basic SEO elements
        - Content optimization
        - Technical implementation
        - User experience
        - Competitive positioning
        - Advanced SEO features
        - Performance metrics
        
        Args:
            url: The URL to analyze
            options: Optional configuration dictionary with the following keys:
                - target_keyword: Focus keyword for analysis
                - competitor_urls: List of competitor URLs
                - depth: Analysis depth ('basic', 'advanced', 'complete')
                - custom_thresholds: Custom analysis thresholds
                - export_format: Desired export format
                - cache_results: Whether to cache results
        
        Returns:
            Dictionary containing comprehensive analysis results
        """
        # Set default options
        default_options = {
            'target_keyword': None,
            'competitor_urls': None,
            'depth': 'complete',
            'custom_thresholds': None,
            'export_format': 'json',
            'cache_results': True
        }
        options = {**default_options, **(options or {})}
        
        # Generate cache key based on options
        cache_key = f"comprehensive_analysis_{url}_{hash(str(options))}"
        
        # Check cache if enabled
        if options['cache_results']:
            cached_result = self.cache.get(cache_key)
            if cached_result:
                self.logger.info(f"Retrieved cached comprehensive analysis for URL: {url}")
                return cached_result

        self.logger.info(f"Starting comprehensive analysis for URL: {url}")
        
        # Initialize analysis dictionary
        analysis = {
            'url': url,
            'analysis_options': options,
            'timestamp': time.time(),
            'analysis_modules': {}
        }
        
        try:
            # Basic SEO Analysis
            analysis['analysis_modules']['basic_seo'] = {
                'meta_tags': self.meta_analyzer.analyze(url),
                'content': self.content_analyzer.analyze(url, options['target_keyword'])
            }
            
            # Modern SEO Features
            analysis['analysis_modules']['modern_seo'] = self.modern_analyzer.analyze(url)
            
            # Competitive Analysis (if competitor URLs provided)
            if options['competitor_urls']:
                analysis['analysis_modules']['competitive'] = self.competitive_analyzer.analyze(
                    url, 
                    options['competitor_urls']
                )
            
            # Advanced Analysis (for 'advanced' and 'complete' depth)
            if options['depth'] in ['advanced', 'complete']:
                analysis['analysis_modules']['advanced'] = self.advanced_analyzer.analyze(url)
            
            # Performance Analysis
            analysis['analysis_modules']['performance'] = self._analyze_performance(url)
            
            # Security Analysis
            analysis['analysis_modules']['security'] = self._analyze_security(url)
            
            # Mobile Optimization
            analysis['analysis_modules']['mobile'] = self._analyze_mobile_optimization(url)
            
            # Generate Insights
            analysis['insights'] = self._generate_comprehensive_insights(analysis)
            
            # Calculate Overall Scores
            analysis['scores'] = self._calculate_comprehensive_scores(analysis)
            
            # Generate Action Plan
            analysis['action_plan'] = self._generate_action_plan(analysis)
            
            # Add Summary
            analysis['summary'] = self._generate_comprehensive_summary(analysis)
            
        except Exception as e:
            self.logger.error(f"Error during comprehensive analysis: {str(e)}")
            analysis['error'] = str(e)
        
        # Cache results if enabled
        if options['cache_results']:
            self.cache.set(cache_key, analysis)
        
        return analysis

    def _analyze_performance(self, url: str) -> Dict:
        """Analyze website performance metrics."""
        return {
            'load_time': self._measure_load_time(url),
            'resource_optimization': self._check_resource_optimization(url),
            'caching_implementation': self._check_caching(url),
            'compression': self._check_compression(url),
            'cdn_usage': self._check_cdn_usage(url)
        }

    def _analyze_security(self, url: str) -> Dict:
        """Analyze website security implementation."""
        return {
            'ssl_certificate': self._check_ssl(url),
            'security_headers': self._check_security_headers(url),
            'content_security_policy': self._check_csp(url),
            'xss_protection': self._check_xss_protection(url),
            'mixed_content': self._check_mixed_content(url)
        }

    def _analyze_mobile_optimization(self, url: str) -> Dict:
        """Analyze mobile optimization implementation."""
        return {
            'viewport_configuration': self._check_viewport(url),
            'responsive_design': self._check_responsive_design(url),
            'touch_elements': self._check_touch_elements(url),
            'mobile_performance': self._check_mobile_performance(url),
            'app_integration': self._check_app_integration(url)
        }

    def _generate_comprehensive_insights(self, analysis: Dict) -> Dict:
        """Generate insights from comprehensive analysis results."""
        insights = {
            'critical_issues': [],
            'major_improvements': [],
            'minor_improvements': [],
            'positive_aspects': [],
            'competitive_edges': [],
            'market_opportunities': []
        }
        
        # Analyze basic SEO elements
        if 'basic_seo' in analysis['analysis_modules']:
            self._analyze_basic_seo_insights(analysis['analysis_modules']['basic_seo'], insights)
        
        # Analyze modern SEO features
        if 'modern_seo' in analysis['analysis_modules']:
            self._analyze_modern_seo_insights(analysis['analysis_modules']['modern_seo'], insights)
        
        # Analyze competitive positioning
        if 'competitive' in analysis['analysis_modules']:
            self._analyze_competitive_insights(analysis['analysis_modules']['competitive'], insights)
        
        # Analyze advanced features
        if 'advanced' in analysis['analysis_modules']:
            self._analyze_advanced_insights(analysis['analysis_modules']['advanced'], insights)
        
        return insights

    def _calculate_comprehensive_scores(self, analysis: Dict) -> Dict:
        """Calculate comprehensive SEO scores."""
        scores = {
            'overall_score': 0,
            'category_scores': {
                'technical_seo': self._calculate_technical_score(analysis),
                'content_quality': self._calculate_content_score(analysis),
                'user_experience': self._calculate_ux_score(analysis),
                'performance': self._calculate_performance_score(analysis),
                'mobile_optimization': self._calculate_mobile_score(analysis),
                'security': self._calculate_security_score(analysis)
            }
        }
        
        # Calculate weighted overall score
        weights = {
            'technical_seo': 0.25,
            'content_quality': 0.25,
            'user_experience': 0.15,
            'performance': 0.15,
            'mobile_optimization': 0.1,
            'security': 0.1
        }
        
        scores['overall_score'] = sum(
            scores['category_scores'][category] * weight
            for category, weight in weights.items()
        )
        
        return scores

    def _generate_action_plan(self, analysis: Dict) -> Dict:
        """Generate prioritized action plan based on analysis results."""
        return {
            'critical_actions': self._get_critical_actions(analysis),
            'high_priority': self._get_high_priority_actions(analysis),
            'medium_priority': self._get_medium_priority_actions(analysis),
            'low_priority': self._get_low_priority_actions(analysis),
            'monitoring_tasks': self._get_monitoring_tasks(analysis)
        }

    def _generate_comprehensive_summary(self, analysis: Dict) -> Dict:
        """Generate executive summary of comprehensive analysis."""
        return {
            'overview': {
                'total_issues': len(analysis['insights']['critical_issues']) + len(analysis['insights']['major_improvements']),
                'critical_issues': len(analysis['insights']['critical_issues']),
                'overall_score': analysis['scores']['overall_score'],
                'strongest_category': max(analysis['scores']['category_scores'].items(), key=lambda x: x[1])[0],
                'weakest_category': min(analysis['scores']['category_scores'].items(), key=lambda x: x[1])[0]
            },
            'key_findings': self._get_key_findings(analysis),
            'competitive_position': self._get_competitive_position(analysis),
            'estimated_impact': self._estimate_improvement_impact(analysis)
        }

    def _measure_load_time(self, url: str) -> Dict:
        """Measure page load time metrics."""
        try:
            start_time = time.time()
            response = requests.get(url, headers=self.settings['crawling']['user_agent'])
            end_time = time.time()
            
            return {
                'total_time': end_time - start_time,
                'time_to_first_byte': response.elapsed.total_seconds(),
                'download_time': (end_time - start_time) - response.elapsed.total_seconds()
            }
        except Exception as e:
            self.logger.error(f"Error measuring load time: {str(e)}")
            return {'error': str(e)}

    def _check_resource_optimization(self, url: str) -> Dict:
        """Check resource optimization implementation."""
        try:
            response = requests.get(url, headers=self.settings['crawling']['user_agent'])
            soup = BeautifulSoup(response.text, 'html.parser')
            
            return {
                'minified_resources': self._check_minification(soup),
                'image_optimization': self._check_image_optimization(soup),
                'resource_hints': self._check_resource_hints(soup),
                'async_loading': self._check_async_loading(soup)
            }
        except Exception as e:
            self.logger.error(f"Error checking resource optimization: {str(e)}")
            return {'error': str(e)}

    def _check_minification(self, soup: BeautifulSoup) -> Dict:
        """Check resource minification status."""
        scripts = soup.find_all('script', src=True)
        styles = soup.find_all('link', rel='stylesheet')
        
        return {
            'js_minified': len([s for s in scripts if '.min.js' in s['src']]) / max(len(scripts), 1),
            'css_minified': len([s for s in styles if '.min.css' in s['href']]) / max(len(styles), 1),
            'total_resources': len(scripts) + len(styles),
            'minification_opportunities': [
                s['src'] for s in scripts if '.min.js' not in s['src']
            ] + [
                s['href'] for s in styles if '.min.css' not in s['href']
            ]
        }

    def _check_image_optimization(self, soup: BeautifulSoup) -> Dict:
        """Check image optimization implementation."""
        images = soup.find_all('img')
        
        return {
            'total_images': len(images),
            'responsive_images': len([img for img in images if img.get('srcset') or img.get('sizes')]),
            'lazy_loading': len([img for img in images if img.get('loading') == 'lazy']),
            'alt_texts': len([img for img in images if img.get('alt')]),
            'optimization_opportunities': [
                img['src'] for img in images 
                if not img.get('srcset') and not img.get('loading') == 'lazy'
            ]
        }

    def _check_resource_hints(self, soup: BeautifulSoup) -> Dict:
        """Check implementation of resource hints."""
        hints = soup.find_all('link', rel=lambda x: x in ['preload', 'prefetch', 'preconnect', 'dns-prefetch'])
        
        return {
            'total_hints': len(hints),
            'by_type': {
                'preload': len([h for h in hints if h['rel'] == ['preload']]),
                'prefetch': len([h for h in hints if h['rel'] == ['prefetch']]),
                'preconnect': len([h for h in hints if h['rel'] == ['preconnect']]),
                'dns-prefetch': len([h for h in hints if h['rel'] == ['dns-prefetch']])
            },
            'optimization_opportunities': bool(len(hints) < 3)
        }

    def _check_async_loading(self, soup: BeautifulSoup) -> Dict:
        """Check asynchronous resource loading."""
        scripts = soup.find_all('script', src=True)
        
        return {
            'total_scripts': len(scripts),
            'async_scripts': len([s for s in scripts if s.get('async')]),
            'defer_scripts': len([s for s in scripts if s.get('defer')]),
            'optimization_opportunities': [
                s['src'] for s in scripts 
                if not s.get('async') and not s.get('defer')
            ]
        }

    def _check_caching(self, url: str) -> Dict:
        # Implementation of _check_caching method
        pass

    def _check_compression(self, url: str) -> Dict:
        # Implementation of _check_compression method
        pass

    def _check_cdn_usage(self, url: str) -> Dict:
        # Implementation of _check_cdn_usage method
        pass

    def _check_ssl(self, url: str) -> Dict:
        # Implementation of _check_ssl method
        pass

    def _check_security_headers(self, url: str) -> Dict:
        # Implementation of _check_security_headers method
        pass

    def _check_csp(self, url: str) -> Dict:
        # Implementation of _check_csp method
        pass

    def _check_xss_protection(self, url: str) -> Dict:
        # Implementation of _check_xss_protection method
        pass

    def _check_mixed_content(self, url: str) -> Dict:
        # Implementation of _check_mixed_content method
        pass

    def _check_viewport(self, url: str) -> Dict:
        # Implementation of _check_viewport method
        pass

    def _check_responsive_design(self, url: str) -> Dict:
        # Implementation of _check_responsive_design method
        pass

    def _check_touch_elements(self, url: str) -> Dict:
        # Implementation of _check_touch_elements method
        pass

    def _check_mobile_performance(self, url: str) -> Dict:
        # Implementation of _check_mobile_performance method
        pass

    def _check_app_integration(self, url: str) -> Dict:
        # Implementation of _check_app_integration method
        pass

    def _analyze_basic_seo_insights(self, basic_seo: Dict, insights: Dict) -> None:
        """Analyze basic SEO elements and generate insights."""
        meta = basic_seo.get('meta_tags', {})
        content = basic_seo.get('content', {})
        
        # Check title and meta description
        if meta.get('title_length', 0) < 30 or meta.get('title_length', 0) > 60:
            insights['major_improvements'].append(
                "Title tag length is not optimal (should be 30-60 characters)"
            )
        else:
            insights['positive_aspects'].append("Title tag length is optimal")
            
        if meta.get('meta_description_length', 0) < 120 or meta.get('meta_description_length', 0) > 160:
            insights['major_improvements'].append(
                "Meta description length is not optimal (should be 120-160 characters)"
            )
        else:
            insights['positive_aspects'].append("Meta description length is optimal")
        
        # Check content quality
        if content.get('word_count', 0) < 300:
            insights['major_improvements'].append(
                "Content length is too short (minimum 300 words recommended)"
            )
        else:
            insights['positive_aspects'].append("Content length is sufficient")
            
        if content.get('keyword_density', 0) > 3.0:
            insights['major_improvements'].append(
                "Keyword density is too high (maximum 3% recommended)"
            )
        else:
            insights['positive_aspects'].append("Keyword density is optimal")

    def _analyze_modern_seo_insights(self, modern_seo: Dict, insights: Dict) -> None:
        """Analyze modern SEO features and generate insights."""
        # Check schema markup
        if not modern_seo.get('has_schema'):
            insights['major_improvements'].append(
                "Missing Schema.org markup for rich results"
            )
        else:
            insights['positive_aspects'].append("Implemented Schema.org markup")
        
        # Check social meta tags
        if not modern_seo.get('has_social_meta'):
            insights['minor_improvements'].append(
                "Missing social media meta tags (Open Graph, Twitter Cards)"
            )
        else:
            insights['positive_aspects'].append("Social media meta tags implemented")
        
        # Check mobile optimization
        if not modern_seo.get('mobile_friendly'):
            insights['critical_issues'].append(
                "Page is not mobile-friendly"
            )
        else:
            insights['positive_aspects'].append("Page is mobile-friendly")

    def _analyze_competitive_insights(self, competitive: Dict, insights: Dict) -> None:
        """Analyze competitive positioning and generate insights."""
        relative_position = competitive.get('relative_position', {})
        
        # Analyze content length compared to competitors
        if relative_position.get('content_length', {}).get('difference', 0) < 0:
            insights['major_improvements'].append(
                f"Content length is below competitor average by {abs(relative_position['content_length']['difference'])} words"
            )
        else:
            insights['competitive_edges'].append(
                f"Content length is above competitor average by {relative_position['content_length']['difference']} words"
            )
        
        # Analyze feature implementation
        features = competitive.get('feature_comparison', {})
        missing_features = [
            feature for feature, implemented in features.items()
            if not implemented and any(comp.get(feature) for comp in competitive.get('competitor_features', []))
        ]
        
        if missing_features:
            insights['market_opportunities'].extend([
                f"Implement {feature.replace('_', ' ')} to match competitors"
                for feature in missing_features
            ])

    def _analyze_advanced_insights(self, advanced: Dict, insights: Dict) -> None:
        """Analyze advanced SEO features and generate insights."""
        # Analyze user experience
        ux = advanced.get('user_experience', {})
        if not ux.get('navigation', {}).get('has_clear_structure'):
            insights['major_improvements'].append(
                "Navigation structure needs improvement for better user experience"
            )
        
        # Analyze accessibility
        acc = ux.get('accessibility', {})
        if acc.get('image_alts', 0) < 0.9:
            insights['major_improvements'].append(
                "Add alt text to images for better accessibility"
            )
        
        # Analyze content clustering
        clusters = advanced.get('content_clusters', {})
        if not clusters.get('topic_hierarchy', {}).get('has_proper_hierarchy'):
            insights['major_improvements'].append(
                "Improve content hierarchy with proper heading structure"
            )
        
        # Analyze rich results potential
        rich = advanced.get('rich_results', {})
        if not rich.get('schema_types'):
            insights['market_opportunities'].append(
                "Implement structured data for rich results in search"
            )

    def _calculate_technical_score(self, analysis_modules: Dict) -> int:
        """Calculate technical SEO score component."""
        score = 0
        
        # HTTPS check (8 points)
        if analysis_modules.get('modern_seo', {}).get('security', {}).get('https'):
            score += 8
        
        # Title tag (7 points)
        title = analysis_modules.get('basic_seo', {}).get('title', '')
        if title and 30 <= len(title) <= 60:
            score += 7
        elif title:
            score += 4
        
        # Meta description (7 points)
        meta_desc = analysis_modules.get('basic_seo', {}).get('meta_description', '')
        if meta_desc and 120 <= len(meta_desc) <= 160:
            score += 7
        elif meta_desc:
            score += 4
        
        # H1 tags (6 points)
        h1_tags = analysis_modules.get('basic_seo', {}).get('headings', {}).get('h1', [])
        if len(h1_tags) == 1:
            score += 6
        elif len(h1_tags) > 0:
            score += 3
        
        # URL structure (4 points)
        url_analysis = analysis_modules.get('advanced_seo', {}).get('url_analysis', {})
        if url_analysis.get('is_seo_friendly'):
            score += 4
        elif url_analysis.get('seo_score', 0) > 50:
            score += 2
        
        # Schema markup (3 points)
        schema_count = len(analysis_modules.get('modern_seo', {}).get('structured_data', {}).get('schema_types', []))
        if schema_count > 0:
            score += min(3, schema_count)
        
        return score
    
    def _calculate_content_score(self, analysis_modules: Dict) -> int:
        """Calculate content quality score component."""
        score = 0
        content_analysis = analysis_modules.get('content', {})
        
        # Word count (12 points)
        word_count = content_analysis.get('basic_metrics', {}).get('word_count', 0)
        if word_count >= 1000:
            score += 12
        elif word_count >= 500:
            score += 9
        elif word_count >= 300:
            score += 6
        elif word_count > 0:
            score += 3
        
        # Readability (10 points)
        flesch_score = content_analysis.get('readability', {}).get('flesch_reading_ease', 0)
        if 60 <= flesch_score <= 80:
            score += 10
        elif 50 <= flesch_score <= 90:
            score += 8
        elif flesch_score > 0:
            score += 4
        
        # Keyword optimization (8 points)
        keyword_analysis = content_analysis.get('keyword_analysis', {})
        if keyword_analysis.get('total_meaningful_words', 0) > 0:
            score += 4
            # Check for keyword stuffing (penalty)
            if len(keyword_analysis.get('keyword_stuffing_detected', [])) == 0:
                score += 4
            else:
                score -= 2  # Penalty for keyword stuffing
        
        # Content structure (5 points)
        structure = content_analysis.get('content_structure', {})
        if structure.get('content_sections', 0) > 2:
            score += 3
        if structure.get('avg_paragraph_length', 0) > 0:
            score += 2
        
        return score
    
    def _calculate_ux_score(self, analysis_modules: Dict) -> int:
        """Calculate user experience score component."""
        score = 0
        
        # Mobile optimization (12 points)
        mobile_analysis = analysis_modules.get('modern_seo', {}).get('mobile_friendly', {})
        if mobile_analysis.get('responsive_viewport'):
            score += 6
        if mobile_analysis.get('viewport_meta'):
            score += 4
        if mobile_analysis.get('touch_elements_spacing', {}).get('potentially_small', 0) == 0:
            score += 2
        
        # Performance indicators (8 points)
        html_structure = analysis_modules.get('modern_seo', {}).get('html_structure', {})
        if html_structure.get('optimizations', {}).get('image_optimization', {}).get('missing_alt', 0) == 0:
            score += 8
        
        return score
    
    def _calculate_security_score(self, analysis_modules: Dict) -> int:
        """Calculate security score component."""
        score = 0
        security_analysis = analysis_modules.get('modern_seo', {}).get('security', {})
        
        if security_analysis.get('https') and security_analysis.get('ssl_certificate_valid'):
            score += 6
        elif security_analysis.get('https'):
            score += 4
        
        if security_analysis.get('hsts'):
            score += 2
        if security_analysis.get('xss_protection'):
            score += 1
        if security_analysis.get('content_security'):
            score += 1
        
        return score
    
    def _get_strongest_category(self, analysis_modules: Dict) -> str:
        """Identify the strongest SEO category."""
        scores = {
            'Technical SEO': self._calculate_technical_score(analysis_modules),
            'Content Quality': self._calculate_content_score(analysis_modules),
            'User Experience': self._calculate_ux_score(analysis_modules),
            'Security': self._calculate_security_score(analysis_modules)
        }
        return max(scores, key=scores.get)
    
    def _get_weakest_category(self, analysis_modules: Dict) -> str:
        """Identify the weakest SEO category."""
        scores = {
            'Technical SEO': self._calculate_technical_score(analysis_modules),
            'Content Quality': self._calculate_content_score(analysis_modules),
            'User Experience': self._calculate_ux_score(analysis_modules),
            'Security': self._calculate_security_score(analysis_modules)
        }
        return min(scores, key=scores.get)

    def _get_critical_actions(self, analysis: Dict) -> List[str]:
        """Get list of critical actions based on analysis."""
        actions = []
        
        # Add critical issues from insights
        actions.extend(analysis.get('insights', {}).get('critical_issues', []))
        
        # Check security issues
        security = analysis['analysis_modules'].get('security', {})
        if not security.get('ssl_certificate', {}).get('is_valid'):
            actions.append("Fix SSL certificate issues")
        
        # Check critical performance issues
        performance = analysis['analysis_modules'].get('performance', {})
        if performance.get('load_time', {}).get('total_time', 0) > 5:
            actions.append("Address critical performance issues - page load time > 5s")
        
        return actions

    def _get_high_priority_actions(self, analysis: Dict) -> List[str]:
        # Implementation of _get_high_priority_actions method
        pass

    def _get_medium_priority_actions(self, analysis: Dict) -> List[str]:
        # Implementation of _get_medium_priority_actions method
        pass

    def _get_low_priority_actions(self, analysis: Dict) -> List[str]:
        # Implementation of _get_low_priority_actions method
        pass

    def _get_monitoring_tasks(self, analysis: Dict) -> List[str]:
        # Implementation of _get_monitoring_tasks method
        pass

    def _get_key_findings(self, analysis: Dict) -> List[str]:
        """Get key findings from analysis results."""
        findings = []
        
        # Add strongest aspects
        findings.extend([
            f"Strength: {aspect}"
            for aspect in analysis.get('insights', {}).get('positive_aspects', [])[:3]
        ])
        
        # Add major improvement areas
        findings.extend([
            f"Improvement: {improvement}"
            for improvement in analysis.get('insights', {}).get('major_improvements', [])[:3]
        ])
        
        # Add competitive insights
        findings.extend([
            f"Competition: {insight}"
            for insight in analysis.get('insights', {}).get('market_opportunities', [])[:2]
        ])
        
        return findings

    def _get_competitive_position(self, analysis: Dict) -> str:
        # Implementation of _get_competitive_position method
        pass

    def _estimate_improvement_impact(self, analysis: Dict) -> Dict:
        """Estimate potential impact of implementing improvements."""
        return {
            'estimated_traffic_increase': self._calculate_traffic_impact(analysis),
            'estimated_conversion_impact': self._calculate_conversion_impact(analysis),
            'implementation_complexity': self._assess_implementation_complexity(analysis),
            'priority_score': self._calculate_priority_score(analysis)
        }

    def _calculate_traffic_impact(self, analysis: Dict) -> float:
        # Implementation of _calculate_traffic_impact method
        pass

    def _calculate_conversion_impact(self, analysis: Dict) -> float:
        # Implementation of _calculate_conversion_impact method
        pass

    def _assess_implementation_complexity(self, analysis: Dict) -> str:
        # Implementation of _assess_implementation_complexity method
        pass

    def _calculate_priority_score(self, analysis: Dict) -> float:
        # Implementation of _calculate_priority_score method
        pass 

    def _extract_meaningful_content(self, soup: BeautifulSoup) -> str:
        """Extract meaningful content text, excluding scripts, styles, and comments."""
        # Remove script and style elements
        for script in soup(["script", "style"]):
            script.decompose()
        
        # Remove comments
        from bs4 import Comment
        comments = soup.findAll(text=lambda text: isinstance(text, Comment))
        for comment in comments:
            comment.extract()
        
        # Get text from content elements only
        content_elements = soup.find_all(['p', 'h1', 'h2', 'h3', 'h4', 'h5', 'h6', 'li', 'td', 'th', 'div', 'span', 'article', 'section'])
        if content_elements:
            text_content = ' '.join([elem.get_text().strip() for elem in content_elements if elem.get_text().strip()])
        else:
            # Fallback to body text if no content elements found
            body = soup.find('body')
            text_content = body.get_text() if body else soup.get_text()
        
        return text_content

    def _count_meaningful_words(self, text: str) -> int:
        """Count meaningful words, excluding common stop patterns."""
        if not text:
            return 0
        
        # Clean the text
        import re
        # Remove extra whitespace and normalize
        text = re.sub(r'\s+', ' ', text.strip())
        
        # Split into words and filter
        words = text.split()
        meaningful_words = []
        
        for word in words:
            # Remove punctuation and convert to lowercase
            clean_word = re.sub(r'[^\w]', '', word.lower())
            
            # Skip if empty, too short, or looks like code/CSS
            if (len(clean_word) >= 2 and 
                not clean_word.isdigit() and 
                not re.match(r'^(px|em|rem|vh|vw|deg|ms|s)$', clean_word) and
                not re.match(r'^(var|function|return|if|else|for|while|class|id)$', clean_word)):
                meaningful_words.append(clean_word)
        
        return len(meaningful_words)