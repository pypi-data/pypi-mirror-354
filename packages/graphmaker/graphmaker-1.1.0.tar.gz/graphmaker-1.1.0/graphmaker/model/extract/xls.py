from pathlib import Path
from typing import Generator
from openpyxl import load_workbook
import pandas as pd
from tqdm import tqdm
import pyarrow.parquet as pq
import pyarrow as pa
import numpy as np
from graphmaker.model.extract._dataprovider import DataProvider


class XLS(DataProvider):
    def __init__(self, file, batch_size=500):
        self.file = file
        self.file_name = Path(file).stem
        self.batch_size = batch_size

    def connect(self):
        self.workbook = load_workbook(self.file, read_only=True)
        return self

    def close(self):
        self.workbook.close()

    def total_rows(self):
        return self.workbook.active.max_row - 1

    def columns(self):
        return [cell.value for cell in next(self.workbook.active.rows)]

    def collect_unique_values(self, column):
        raise NotImplementedError("This method is not implemented yet.")

    def next(self) -> Generator[list[str | int | float], None, None]:
        sheet = self.workbook.active
        total_rows = sheet.max_row - 1  # Exclude header row

        # Get column names from first row
        columns = [cell.value for cell in next(sheet.rows)]

        # Process rows in batches
        rows = [None] * self.batch_size
        for idx, row in enumerate(sheet.iter_rows(min_row=2, values_only=True)):
            rows[idx % self.batch_size] = row
            if idx == self.batch_size - 1 or idx == total_rows - 1:
                yield rows

    def to_parquet_file(self, output_folder:Path|str, verbose=False):
        if isinstance(output_folder, str):
            output_folder = Path(output_folder)

        try:
            sheets = self.workbook.sheetnames

            for sheet_name in sheets:
                sheet = self.workbook[sheet_name]

                dest = output_folder / f"{self.file_name}".replace(".xlsx", "") / f"{sheet_name}.parquet"

                # Create the directory if it does not exist
                dest.parent.mkdir(parents=True, exist_ok=True)

                columns = [cell.value for cell in next(sheet.rows)]
                total_rows = sheet.max_row - 1

                print(columns)
                columns = [col if col is not None else "" for col in columns]
                # maybe we can make this better by checking the type of the column
                schema = pa.schema([pa.field(col, pa.string()) for col in columns])

                records = np.empty((self.batch_size, len(columns)), dtype=object)

                with pq.ParquetWriter(dest, schema=schema) as writer:
                    iterator = (
                        tqdm(
                            sheet.iter_rows(values_only=True, min_row=2),
                            total=sheet.max_row - 1,
                            desc="Parquet",
                        )
                        if verbose
                        else sheet.iter_rows(values_only=True, min_row=2)
                    )
                    for i, r in enumerate(iterator):
                        records[i % self.batch_size] = np.asarray(r)
                        if (i + 1) % self.batch_size == 0 or i + 1 == total_rows:
                            if i + 1 == total_rows:
                                records = records[0: total_rows % self.batch_size]
                            batch = pa.RecordBatch.from_arrays(records.astype(str).T.tolist(), schema=schema)
                            writer.write_batch(batch)

                # try to open the file, so that it fails if the file is corrupted
                df = pd.read_parquet(dest, engine="pyarrow")
                assert df.shape[0] == total_rows, f"Error: {df.shape[0]} != {total_rows}"
                if verbose:
                    print(f"Data written to {dest} in Parquet format.")
                return dest
        except FileNotFoundError:
            print(f"File {self.file} not found.")
            raise FileNotFoundError(f"File {self.file} not found.")
        except PermissionError:
            print(f"Permission denied for file {self.file}.")
            raise PermissionError(f"Permission denied for file {self.file}.")
        except OSError as e:
            print(f"OS error: {e}")
            raise OSError(f"OS error: {e}")
        except Exception as e:
            print(f"Error processing file {self.file}: {e}")
            raise e