import pandas as pd
import os
from openpyxl import load_workbook



class Utilities:
    """Parent utility class containing various utility subclasses."""

    class ExcelUtils:
        """Excel utility class containing various Excel utilities."""

        @staticmethod
        def create_excel(file_path: str, columns: list):
            """Create a new Excel file with given columns."""
            df = pd.DataFrame(columns=columns)
            df.to_excel(file_path, index=False)

        @staticmethod
        def append_row(file_path: str, data: dict):
            """Append a row of data to an existing Excel file."""
            new_row = pd.DataFrame([data])
            if not os.path.exists(file_path):
                new_row.to_excel(file_path, index=False)
            else:
                with pd.ExcelWriter(file_path, mode="a", if_sheet_exists="overlay", engine="openpyxl") as writer:
                    sheet = load_workbook(file_path).active
                    startrow = sheet.max_row
                    new_row.to_excel(writer, index=False, header=False, startrow=startrow)

        @staticmethod
        def read_excel(file_path: str) -> pd.DataFrame:
            """Read the Excel file into a DataFrame."""
            return pd.read_excel(file_path)

        @staticmethod
        def filter_rows(file_path: str, column: str, value) -> pd.DataFrame:
            """Return filtered DataFrame where column == value."""
            df = pd.read_excel(file_path)
            return df[df[column] == value]

        @staticmethod
        def update_cell(file_path: str, row_index: int, column_name: str, new_value):
            """Update a specific cell in the Excel file."""
            df = pd.read_excel(file_path)
            df.at[row_index, column_name] = new_value
            df.to_excel(file_path, index=False)


        @staticmethod
        def delete_row(file_path: str, row_index: int):
            """Delete a row by index from the Excel file."""
            df = pd.read_excel(file_path)
            df.drop(index=row_index, inplace=True)
            df.reset_index(drop=True, inplace=True)
            df.to_excel(file_path, index=False)

        @staticmethod
        def get_row_count(file_path: str) -> int:
            """Return the number of rows (excluding header) in the Excel file."""
            df = pd.read_excel(file_path)
            return len(df)

        @staticmethod
        def get_column_names(file_path: str) -> list:
            """Return the list of column names."""
            df = pd.read_excel(file_path)
            return df.columns.tolist()


        @staticmethod
        def sort_by_column(file_path: str, column: str, ascending=True) -> pd.DataFrame:
            """Sort the DataFrame by a given column."""
            df = pd.read_excel(file_path)
            df_sorted = df.sort_values(by=column, ascending=ascending)
            df_sorted.to_excel(file_path, index=False)
            return df_sorted

        @staticmethod
        def rename_column(file_path: str, old_name: str, new_name: str):
            """Rename a column in the Excel file."""
            df = pd.read_excel(file_path)
            df.rename(columns={old_name: new_name}, inplace=True)
            df.to_excel(file_path, index=False)

        @staticmethod
        def clear_excel(file_path: str, keep_header=True):
            """Clear all data (or everything) from the Excel file."""
            df = pd.read_excel(file_path)
            if keep_header:
                df.iloc[0:0].to_excel(file_path, index=False)
            else:
                pd.DataFrame().to_excel(file_path, index=False)

        @staticmethod
        def get_unique_values(file_path: str, column: str) -> list:
            """Return list of unique values in a specific column."""
            df = pd.read_excel(file_path)
            return df[column].dropna().unique().tolist()

        @staticmethod
        def column_stats(file_path: str, column: str) -> dict:
            """Return basic stats (min, max, mean, median) for numeric column."""
            df = pd.read_excel(file_path)
            if not pd.api.types.is_numeric_dtype(df[column]):
                raise ValueError("Selected column is not numeric.")
            return {
                "min": df[column].min(),
                "max": df[column].max(),
                "mean": df[column].mean(),
                "median": df[column].median(),
            }
        
        @staticmethod
        def highlight_negative(file_path: str, column: str):
            """Highlight negative numbers in red background in the given column.

            Example:
                Utilities.ExcelUtils.highlight_negative("data.xlsx", "B")
            """
            wb = load_workbook(file_path)
            ws = wb.active
            col_idx = ws[column + "1"].column
            for row in ws.iter_rows(min_row=2):  # Skip header
                cell = row[col_idx - 1]
                if isinstance(cell.value, (int, float)) and cell.value < 0:
                    cell.fill = PatternFill(start_color="FFC7CE", fill_type="solid")
            wb.save(file_path)

        @staticmethod
        def list_sheets(file_path: str) -> list:
            """List all sheet names in the Excel file.

            Example:
                sheets = Utilities.ExcelUtils.list_sheets("data.xlsx")
            """
            wb = load_workbook(file_path)
            return wb.sheetnames

        @staticmethod
        def transpose(file_path: str, output_path: Optional[str] = None):
            """Transpose Excel data (rows <-> columns).

            Example:
                Utilities.ExcelUtils.transpose("data.xlsx", "transposed.xlsx")
            """
            df = pd.read_excel(file_path)
            df.T.to_excel(output_path or file_path, header=False)

        @staticmethod
        def add_dropdown(file_path: str, column_letter: str, options: list, start_row: int = 2):
            """Add a dropdown list to cells in a column.

            Example:
                Utilities.ExcelUtils.add_dropdown("file.xlsx", "C", ["Yes", "No"])
            """
            wb = load_workbook(file_path)
            ws = wb.active
            dv = DataValidation(type="list", formula1=f'"{','.join(options)}"', showDropDown=True)
            ws.add_data_validation(dv)
            dv.add(f"{column_letter}{start_row}:{column_letter}1048576")
            wb.save(file_path)

        @staticmethod
        def find_duplicates(file_path: str, column: Optional[str] = None) -> pd.DataFrame:
            """Return duplicated rows or values in a column.

            Example:
                df = Utilities.ExcelUtils.find_duplicates("file.xlsx", "Name")
            """
            df = pd.read_excel(file_path)
            if column:
                return df[df.duplicated(subset=column, keep=False)]
            return df[df.duplicated(keep=False)]

        @staticmethod
        def protect_excel(file_path: str, password: str):
            """Protect Excel file with password (requires Excel on system).

            Example:
                Utilities.ExcelUtils.protect_excel("file.xlsx", "1234")
            """
            try:
                import win32com.client
                excel = win32com.client.Dispatch("Excel.Application")
                wb = excel.Workbooks.Open(file_path)
                wb.Password = password
                wb.Save()
                wb.Close()
                excel.Quit()
            except ImportError:
                raise ImportError("win32com module is required. Only works on Windows with Excel installed.")

        @staticmethod
        def autofit_columns(file_path: str):
            """Auto-adjust column widths to fit content.

            Example:
                Utilities.ExcelUtils.autofit_columns("data.xlsx")
            """
            wb = load_workbook(file_path)
            ws = wb.active
            for col in ws.columns:
                max_length = max(len(str(cell.value or "")) for cell in col)
                col_letter = col[0].column_letter
                ws.column_dimensions[col_letter].width = max_length + 2
            wb.save(file_path)

