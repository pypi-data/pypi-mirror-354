# Copyright 2025 Ravetta Stefano
#
# Licensed under the Apache License, Version 2.0 (the "License");
# you may not use this file except in compliance with the License.
# You may obtain a copy of the License at
#
#     http://www.apache.org/licenses/LICENSE-2.0
#
# Unless required by applicable law or agreed to in writing, software
# distributed under the License is distributed on an "AS IS" BASIS,
# WITHOUT WARRANTIES OR CONDITIONS OF ANY KIND, either express or implied.
# See the License for the specific language governing permissions and
# limitations under the License.

from typing import Dict, List

from openpyxl.workbook import Workbook
from openpyxl.worksheet.worksheet import Worksheet


class ExcelWriter:
    def __init__(self, filename):
        self._filename = filename
        self._workbook: Workbook = Workbook()
        self._worksheet: Dict[str, Worksheet] = {}

    def create_sheet(self, sheet_name: str):
        if len(self._worksheet.keys()) == 0:
            self._worksheet[sheet_name] = self._workbook.active
            self._worksheet[sheet_name].title = sheet_name
        else:
            self._worksheet[sheet_name] = self._workbook.create_sheet(title=sheet_name)

    def append_in_sheet(self, sheet_name: str, data: List):
        self._worksheet[sheet_name].append(data)

    def write(self):
        self._workbook.save(self._filename)
