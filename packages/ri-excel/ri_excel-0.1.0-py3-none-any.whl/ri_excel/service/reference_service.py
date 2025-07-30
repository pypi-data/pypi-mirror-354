# Copyright 2025 Ravetta Stefano
#
# Licensed under the Apache License, Version 2.0 (the "License");
# you may not use this file except in compliance with the License.
# You may obtain a copy of the License at
#
#     http://www.apache.org/licenses/LICENSE-2.0
#
# Unless required by applicable law or agreed to in writing, software
# distributed under the License is distributed on an "AS IS" BASIS,
# WITHOUT WARRANTIES OR CONDITIONS OF ANY KIND, either express or implied.
# See the License for the specific language governing permissions and
# limitations under the License.

from typing import Dict, List, Any

from openpyxl.worksheet.worksheet import Worksheet

from ri_excel.model.cell_config import Cell, CellConfig, CellType
from ri_excel.model.reference_config import ReferenceConfig, ReferenceType
from ri_excel.util import util


class ReferenceService:

    def __init__(self, config: ReferenceConfig, worksheet: Worksheet):
        self._config: ReferenceConfig = config
        self._worksheet: Worksheet = worksheet
        self._reference_values: Dict[str, Cell | List[Cell]] = {}

    def load_reference_values(self):
        if self._config.type == ReferenceType.START_FIXED_TO_END_OF_DATA:
            self._start_fixed_to_end_of_data_load_values()

        else:
            self._raise_with_invalid_type()

    def iter_rows(self):
        if self._config.type == ReferenceType.START_FIXED_TO_END_OF_DATA:
            return self._start_fixed_to_end_of_data_iter_rows()

        else:
            self._raise_with_invalid_type()

    def get_cell_value(self, cell: CellConfig, key_value: str | None = None) -> Any:
        if cell.type == CellType.VALUE:
            return self._worksheet.cell(cell.row, cell.col).value

        elif cell.type == CellType.REFERENCE:
            if key_value not in self._reference_values:
                return None

            reference_value = self._reference_values[key_value]

            if self._config.duplicated:
                # TODO: return the list of cell data based on the references, for each reference, replace the None
                #       coordinate with the coordinate of the referenced cell

                pass

            # Retrieves the cell data based on the reference and replaces the None coordinate with the coordinate
            # of the referenced cell
            col = cell.col
            row = cell.row

            if (col is None and row is None) or (col is not None and row is not None):
                raise Exception(f'Cell of type {cell.type} must have exactly one coordinate set')

            elif col is None:
                col = reference_value.col

                # TODO: Slow operation, figure out how to optimize in the future (useful only for fairly unusual
                #       horizontal tables)
                return self._worksheet.cell(row, col).value

            elif row is None:
                return reference_value.row_values[col - 1].value

            else:
                return None

        else:
            raise Exception(f'Cell type {cell.type} is not valid')

    def _start_fixed_to_end_of_data_load_values(self):
        min_row = self._config.cell.row
        col = self._config.cell.col

        row = min_row
        for row_values in self._worksheet.iter_rows(min_row, self._worksheet.max_row, 1, self._worksheet.max_column):
            value = row_values[col - 1].value

            if util.regex_match(self._config.regex, value):
                if value not in self._reference_values:
                    if self._config.duplicated:
                        self._reference_values[value] = [Cell(col=col, row=row, row_values=row_values)]
                    else:
                        self._reference_values[value] = Cell(col=col, row=row, row_values=row_values)

                elif self._config.duplicated:
                    self._reference_values[value].append(Cell(col=col, row=row, row_values=row_values))

            row += 1

    def _start_fixed_to_end_of_data_iter_rows(self):
        min_row = self._config.cell.row

        for row_values in self._worksheet.iter_rows(min_row, self._worksheet.max_row, 1, self._worksheet.max_column):
            if util.regex_match(self._config.regex, row_values[self._config.cell.col - 1].value):
                yield row_values

    def _raise_with_invalid_type(self):
        if self._config.type is None or self._config.type == '':
            raise Exception('The type must be set')

        else:
            raise Exception(f"Type '{self._config.type}' is not valid")

