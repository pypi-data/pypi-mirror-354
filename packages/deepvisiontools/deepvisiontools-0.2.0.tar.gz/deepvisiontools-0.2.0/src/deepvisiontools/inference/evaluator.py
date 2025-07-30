from deepvisiontools import DeepVisionDataset, Configuration
from deepvisiontools.inference import Predictor
from deepvisiontools.utils import visualization
from tqdm import tqdm
from pathlib import Path
import pandas as pd
from typing import Literal, Union
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill
import random


class Evaluator:
    """Evaluator class : evaluate a given Predictor (model + patch_size + additional Configurations) on a dataset with given metrics.
    The results are saved in a generated csv file (metrics at dataset level) and in a xlsx file (metrics at dataset and sample levels).
    Highlighting samples that deviate from the mean or median (giving deviation_method) by nb_sigma sigma in the xlsx file.
    Create visualizations of predictions giving number_visu.
    Returns the dictionnary with metrics at dataset level and the dictionnary with metrics at sample level.
    Prints a dataframe with metrics at dataset level (same content as in the generated csv file).

    Args:
        predictor (``Predictor``): Predictor class to evaluate
        metrics (``list``): List of metrics to evaluate the Predictor on
        deviation_method (``Literal[&quot;mean&quot;, &quot;median&quot;]``, **optional**): method to compute outlayers. Defaults to "mean".
        nb_sigma (``Union[int, float]``, **optional**): number of standard deviations for outlayers. Defaults to 2.

    Example:
    ----------

    .. highlight:: python
    .. code-block:: python

        >>> from deepvisiontools import Evaluator, Predictor
        >>> from deepvisiontools.metrics import DetectF1Score
        >>> predictor = Predictor(model=\path\to\model.pth)
        >>> evaluator = Evaluator(predictor, metrics=[DetectF1Score()])
        >>> evaluator.evaluate(mydataset, "results")

    Attributes
    ----------

    Attributes:
        predictor (``Predictor``)
        metrics (``List[BaseMetric]``)
        nb_sigma (``int``)
        data_type (``Literal[&quot;instance_mask&quot;, &quot;bbox&quot;, &quot;keypoint&quot;, &quot;semantic_mask&quot;]``)
        deviation_method (```Literal[&quot;mean&quot;, &quot;median&quot;]``)


    **Methods**

    """

    def __init__(
        self,
        predictor: Predictor,
        metrics: list,
        deviation_method: Literal["mean", "median"] = "mean",
        nb_sigma: Union[int, float] = 2,
    ):

        self.predictor = predictor
        self.metrics = metrics
        self.nb_sigma = nb_sigma
        self.data_type = Configuration().data_type
        self.deviation_method = deviation_method
        assert deviation_method in [
            "mean",
            "median",
        ], "deviation_method must be 'mean' or 'median'"

    def evaluate(
        self,
        dataset: DeepVisionDataset,
        result_folder: Union[str, Path],
        number_visu: Union[Literal["all"], int] = "all",
    ):
        """Run evaluation on dataset. Compute metrics for dataset and for each sample of the dataset."""
        result_folder = (
            result_folder if isinstance(result_folder, Path) else Path(result_folder)
        )
        result_folder.mkdir(parents=True, exist_ok=True)
        export_csv_path = result_folder / "results.csv"
        export_xlsx_path = result_folder / "results.xlsx"
        visu_folder_path = result_folder / "visu"

        assert (
            isinstance(number_visu, int) or number_visu == "all"
        ), "number_visu must be integer or equal to 'all'"

        dataset.preprocessing = None  # because already preprocess in Predictor

        # Creation and initialisation of the dictionary containing the metrics for each image (one dictionary/metric containing one dictionary/image)
        metric_sample_dict = {}
        for metric in self.metrics:
            metric.to(Configuration().device)
            metric_sample_dict.update({metric.name: {}})

        if (
            number_visu != 0 and number_visu != "all"
        ):  # randomly select index of the number_visu images to visualize
            list_index_img = [i for i in range(len(dataset))]
            random.shuffle(list_index_img)
            list_index_img_visu = list_index_img[:number_visu]

        # For each sample in the dataset : prediction then calculation and storage of tp, fp, tn, fn (for each metric)
        iterator = tqdm(dataset, total=len(dataset), desc=f"Evaluation progress ")

        for index_img, (image, target, img_name) in enumerate(iterator):
            prediction = self.predictor.predict(image)

            for metric in self.metrics:
                metric.update(
                    prediction, target
                )  # update 'stats' attribute for each metric

                metric_sample_dict[metric.name].update(
                    {img_name: {"img_index_in_dataset": index_img}}
                )  # creation of the image dictionary (for the given metric) with its index information
                # compute metric for the image (all classes combined and for each class)
                metric_sample = metric.compute_last_sample()
                metric_sample_dict[metric.name][img_name].update(metric_sample)

            # prediction visualization
            if number_visu != 0 and number_visu != "all":
                if index_img in list_index_img_visu:
                    visualization(
                        image,
                        prediction,
                        save_path=visu_folder_path
                        / "visu_img{0}_{1}".format(index_img, img_name),
                    )
            elif number_visu == "all":
                visualization(
                    image,
                    prediction,
                    save_path=visu_folder_path
                    / "visu_img{0}_{1}".format(index_img, img_name),
                )

        # after all updates recompute to get averaged values of metric (global, samplewise, macro, micro), save in metric_dict
        metric_dict = {}
        for metric in self.metrics:
            results = metric.compute()
            metric_dict.update({metric.name: results})

        # Converts dictionary metric values from tensor to float :
        for metric_name in metric_dict:
            for sous_metric_name in metric_dict[metric_name]:
                metric_dict[metric_name][sous_metric_name] = round(
                    metric_dict[metric_name][sous_metric_name].item(), 3
                )

        # Converts the dict into a dataframe, which is saved in a csv file:
        df_metrics = pd.DataFrame(metric_dict)
        df_metrics.to_csv(export_csv_path)

        # Create the Excel file with metrics values for each image
        sheet_names = self._create_xls(export_xlsx_path, df_metrics, metric_sample_dict)

        # Modifies the Excel file with metrics : highlights extreme values and adds mean, std, median
        self._formate_xls(export_xlsx_path, sheet_names, metric_sample_dict)

        print(df_metrics)
        return (metric_dict, metric_sample_dict)

    def _create_xls(self, export_xlsx_path, df_metrics, metric_sample_dict):
        """Create an Excel file with metrics values for each image"""
        sheet_names = []
        with pd.ExcelWriter(export_xlsx_path) as writer:
            df_metrics.to_excel(
                writer, sheet_name="Dataset_Metrics", index=True
            )  # creates a first sheet with the 'sub-metrics' for each metric (equivalent to the contents of the csv generated earlier)
            for metric in metric_sample_dict:
                df_sample = pd.DataFrame(metric_sample_dict[metric]).T
                df_sample["Outlier?"] = None
                sheet_name = "{0}_by_samples".format(metric)
                if len(sheet_name) > 31:  # the length of excel sheets name is 31
                    sheet_name = sheet_name[: len(sheet_name) - (len(sheet_name) - 31)]
                df_sample.to_excel(
                    writer, sheet_name=sheet_name, index=True
                )  # creates a sheet by metric with details by image
                sheet_names.append(sheet_name)
        return sheet_names

    def _highlight_cells(self, sheet, colonne_verif, min_val, max_val):
        """Applies a red background and bold text if the value is outside the range"""
        # Style for formatting (red fill + bold text)
        red_fill = PatternFill(
            start_color="FFC7CE", end_color="FFC7CE", fill_type="solid"
        )
        bold_font = Font(bold=True, color="9C0006")

        for row in sheet.iter_rows(
            min_row=2, min_col=colonne_verif, max_col=colonne_verif
        ):
            for cell in row:
                if cell.value is not None and (
                    cell.value < min_val or cell.value > max_val
                ):
                    cell.fill = red_fill
                    sheet.cell(row=cell.row, column=1).font = (
                        bold_font  # Make the index bold and red
                    )
                    sheet.cell(cell.row, sheet.max_column).value = (
                        "True"  # True in "Outlier?" column
                    )

    def _formate_xls(self, export_xlsx_path, sheet_names, metric_sample_dict):
        """Modifies the Excel file with metrics : highlights extreme values and adds mean, std, median"""

        wb = load_workbook(export_xlsx_path)

        for sh_name, metric in zip(sheet_names, metric_sample_dict):

            df_sample = pd.DataFrame(metric_sample_dict[metric]).T
            columns_list = df_sample.columns.tolist()
            sheet_maw_row = wb[sh_name].max_row

            # add mean, std, median labels in last rows
            wb[sh_name].cell(sheet_maw_row + 2, 1).value = "Mean : "
            wb[sh_name].cell(sheet_maw_row + 3, 1).value = "Std : "
            wb[sh_name].cell(sheet_maw_row + 4, 1).value = "Median : "

            # add mean, std and median values in last rows for each class and "all_class"
            for num_col, col_name in enumerate(columns_list[1:]):
                mean = df_sample.mean()[col_name]
                std = df_sample.std()[col_name]
                median = df_sample.median()[col_name]

                num_col_xsl = num_col + 3
                # highlights extreme values
                if self.deviation_method == "median":
                    self._highlight_cells(
                        wb[sh_name],
                        colonne_verif=num_col_xsl,
                        min_val=median - (self.nb_sigma * std),
                        max_val=median + (self.nb_sigma * std),
                    )
                else:
                    self._highlight_cells(
                        wb[sh_name],
                        colonne_verif=num_col_xsl,
                        min_val=mean - (self.nb_sigma * std),
                        max_val=mean + (self.nb_sigma * std),
                    )

                # add mean, std and median in last rows
                wb[sh_name].cell(sheet_maw_row + 2, num_col_xsl).value = round(mean, 3)
                wb[sh_name].cell(sheet_maw_row + 3, num_col_xsl).value = round(std, 3)
                wb[sh_name].cell(sheet_maw_row + 4, num_col_xsl).value = round(
                    median, 3
                )

        # Save file after modifications
        wb.save(export_xlsx_path)
