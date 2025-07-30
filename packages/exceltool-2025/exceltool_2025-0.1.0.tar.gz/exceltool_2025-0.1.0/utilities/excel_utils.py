import pandas as pd
import os
from openpyxl import load_workbook



class Utilities:
    """Parent utility class containing various utility subclasses."""

    class ExcelUtils:
        """Excel utility class containing various Excel utilities."""

        @staticmethod
        def create_excel(file_path: str, columns: list):
            """Create a new Excel file with given columns."""
            df = pd.DataFrame(columns=columns)
            df.to_excel(file_path, index=False)

        @staticmethod
        def append_row(file_path: str, data: dict):
            """Append a row of data to an existing Excel file."""
            new_row = pd.DataFrame([data])
            if not os.path.exists(file_path):
                new_row.to_excel(file_path, index=False)
            else:
                with pd.ExcelWriter(file_path, mode="a", if_sheet_exists="overlay", engine="openpyxl") as writer:
                    sheet = load_workbook(file_path).active
                    startrow = sheet.max_row
                    new_row.to_excel(writer, index=False, header=False, startrow=startrow)

        @staticmethod
        def read_excel(file_path: str) -> pd.DataFrame:
            """Read the Excel file into a DataFrame."""
            return pd.read_excel(file_path)

        @staticmethod
        def filter_rows(file_path: str, column: str, value) -> pd.DataFrame:
            """Return filtered DataFrame where column == value."""
            df = pd.read_excel(file_path)
            return df[df[column] == value]

        @staticmethod
        def update_cell(file_path: str, row_index: int, column_name: str, new_value):
            """Update a specific cell in the Excel file."""
            df = pd.read_excel(file_path)
            df.at[row_index, column_name] = new_value
            df.to_excel(file_path, index=False)


        @staticmethod
        def delete_row(file_path: str, row_index: int):
            """Delete a row by index from the Excel file."""
            df = pd.read_excel(file_path)
            df.drop(index=row_index, inplace=True)
            df.reset_index(drop=True, inplace=True)
            df.to_excel(file_path, index=False)

        @staticmethod
        def get_row_count(file_path: str) -> int:
            """Return the number of rows (excluding header) in the Excel file."""
            df = pd.read_excel(file_path)
            return len(df)

        @staticmethod
        def get_column_names(file_path: str) -> list:
            """Return the list of column names."""
            df = pd.read_excel(file_path)
            return df.columns.tolist()


        @staticmethod
        def sort_by_column(file_path: str, column: str, ascending=True) -> pd.DataFrame:
            """Sort the DataFrame by a given column."""
            df = pd.read_excel(file_path)
            df_sorted = df.sort_values(by=column, ascending=ascending)
            df_sorted.to_excel(file_path, index=False)
            return df_sorted

        @staticmethod
        def rename_column(file_path: str, old_name: str, new_name: str):
            """Rename a column in the Excel file."""
            df = pd.read_excel(file_path)
            df.rename(columns={old_name: new_name}, inplace=True)
            df.to_excel(file_path, index=False)

        @staticmethod
        def clear_excel(file_path: str, keep_header=True):
            """Clear all data (or everything) from the Excel file."""
            df = pd.read_excel(file_path)
            if keep_header:
                df.iloc[0:0].to_excel(file_path, index=False)
            else:
                pd.DataFrame().to_excel(file_path, index=False)

        @staticmethod
        def get_unique_values(file_path: str, column: str) -> list:
            """Return list of unique values in a specific column."""
            df = pd.read_excel(file_path)
            return df[column].dropna().unique().tolist()

        @staticmethod
        def column_stats(file_path: str, column: str) -> dict:
            """Return basic stats (min, max, mean, median) for numeric column."""
            df = pd.read_excel(file_path)
            if not pd.api.types.is_numeric_dtype(df[column]):
                raise ValueError("Selected column is not numeric.")
            return {
                "min": df[column].min(),
                "max": df[column].max(),
                "mean": df[column].mean(),
                "median": df[column].median(),
            }
        

