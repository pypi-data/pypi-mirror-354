"""
JSON转Excel核心转换器
"""

import json
import traceback
from typing import Any, Dict, List
import pandas as pd
from openpyxl.styles import Alignment, Font, PatternFill, Border, Side

from .config import CONVERTER_CONFIG, VALIDATION_CONFIG
from .formatters import CellFormatter


class JsonToExcelConverter:
    """嵌套JSON转Excel转换器"""

    def __init__(self, config: Dict[str, Any] = None):
        """
        初始化转换器

        Args:
            config: 自定义配置，会与默认配置合并
        """
        self.config = {**CONVERTER_CONFIG}
        if config:
            self.config.update(config)

        self.validation_config = VALIDATION_CONFIG

    def flatten_dict(
        self,
        data: Dict[str, Any],
        parent_key: str = "",
        sep: str = None,
        depth: int = 0,
    ) -> Dict[str, Any]:
        """
        递归扁平化嵌套字典

        Args:
            data: 要处理的字典
            parent_key: 父级键名
            sep: 分隔符
            depth: 当前递归深度

        Returns:
            处理后的字典
        """
        if sep is None:
            sep = self.config["separator"]

        if depth > self.config["max_depth"]:
            # 防止无限递归
            return {parent_key: json.dumps(data, ensure_ascii=False)}

        items = []

        for key, value in data.items():
            # 构建新的键名
            new_key = f"{parent_key}{sep}{key}" if parent_key else key

            if isinstance(value, dict):
                # 递归处理嵌套字典
                nested_items = self.flatten_dict(value, new_key, sep, depth + 1)
                items.extend(nested_items.items())
            else:
                # 使用格式化器处理值
                formatted_value = CellFormatter.format_value(value)
                items.append((new_key, formatted_value))

        return dict(items)

    def process_data(self, data: List[Dict[str, Any]], verbose: bool = True) -> List[Dict[str, Any]]:
        """
        批量处理数据列表

        Args:
            data: 嵌套JSON数据列表
            verbose: 是否显示处理进度

        Returns:
            扁平化后的数据列表
        """
        if not data:
            raise ValueError("输入数据为空")

        if len(data) > self.validation_config["max_records"]:
            raise ValueError(
                f"数据量超过限制: {len(data)} > {self.validation_config['max_records']}"
            )

        flattened_data = []

        for i, item in enumerate(data):
            try:
                if isinstance(item, dict):
                    if verbose:
                        print(f"🔄 正在处理第 {i+1} 条记录...")
                    flattened_item = self.flatten_dict(item)
                    if verbose:
                        print(f"✅ 第 {i+1} 条记录处理完成，生成 {len(flattened_item)} 个字段")
                    flattened_data.append(flattened_item)
                else:
                    # 非字典类型处理
                    flattened_data.append(
                        {
                            "record_index": i + 1,
                            "value": CellFormatter.format_value(item),
                        }
                    )
            except Exception as e:
                if verbose:
                    print(f"❌ 处理第 {i+1} 条记录时出错: {str(e)}")
                    traceback.print_exc()
                if self.validation_config["empty_data_handling"] == "error":
                    raise
                elif self.validation_config["empty_data_handling"] == "warning":
                    continue

        if verbose:
            print(f"📊 数据处理完成，共处理 {len(flattened_data)} 条记录")
        return flattened_data

    def _detect_numeric_columns(self, worksheet, threshold=0.7):
        """
        检测主要包含数字的列
        
        Args:
            worksheet: openpyxl工作表对象
            threshold: 数字比例阈值，默认70%以上为数字则认为是数字列
        
        Returns:
            set: 数字列的列字母集合
        """
        numeric_columns = set()
        
        # 跳过标题行，从第2行开始检查
        if worksheet.max_row <= 1:
            return numeric_columns
        
        for col_idx in range(1, worksheet.max_column + 1):
            column_letter = worksheet.cell(row=1, column=col_idx).column_letter
            
            # 收集该列的所有非空值
            values = []
            for row_idx in range(2, worksheet.max_row + 1):
                cell_value = worksheet.cell(row=row_idx, column=col_idx).value
                if cell_value is not None and str(cell_value).strip():
                    values.append(str(cell_value).strip())
            
            if not values:
                continue
                
            # 检查数字比例
            numeric_count = 0
            for value in values:
                if self._is_numeric_value(value):
                    numeric_count += 1
            
            # 如果数字比例超过阈值，标记为数字列
            if len(values) > 0 and (numeric_count / len(values)) >= threshold:
                numeric_columns.add(column_letter)
        
        return numeric_columns

    def _is_numeric_value(self, value):
        """
        判断值是否为数字（包括整数、浮点数、百分比等）
        
        Args:
            value: 要检查的值
            
        Returns:
            bool: 是否为数字
        """
        try:
            # 移除常见的非数字字符进行检查
            cleaned_value = str(value).replace(',', '').replace('%', '').replace('$', '').replace('￥', '').strip()
            
            # 尝试转换为浮点数
            float(cleaned_value)
            return True
        except (ValueError, TypeError):
            return False

    def _auto_adjust_columns(self, worksheet):
        """自动调整列宽"""
        for column in worksheet.columns:
            max_length = 0
            column_letter = column[0].column_letter

            for cell in column:
                try:
                    cell_value = str(cell.value) if cell.value is not None else ""

                    # 处理换行符
                    if "\n" in cell_value:
                        lines = cell_value.split("\n")
                        max_line_length = (
                            max(len(line) for line in lines) if lines else 0
                        )
                        max_length = max(max_length, max_line_length)
                    else:
                        max_length = max(max_length, len(cell_value))
                except Exception:
                    pass

            # 根据列名长度设置最小列宽
            header_length = len(
                str(worksheet.cell(row=1, column=column[0].column).value or "")
            )
            adjusted_width = min(
                max(
                    max_length + 6,
                    header_length + 4,
                    self.config["min_column_width"],
                ),
                self.config["max_column_width"],
            )

            worksheet.column_dimensions[column_letter].width = adjusted_width

    def _apply_text_wrapping(self, worksheet, numeric_columns=None):
        """
        应用文本换行（修改版本，考虑对齐方式）
        
        Args:
            worksheet: openpyxl工作表对象
            numeric_columns: 数字列集合
        """
        if numeric_columns is None:
            numeric_columns = set()
        
        # 设置标题行高度为28.8
        worksheet.row_dimensions[1].height = 28.8

        # 应用标题行的自动换行（居中）
        for cell in worksheet[1]:
            cell.alignment = Alignment(wrap_text=True, horizontal="center", vertical="center")

        for row in worksheet.iter_rows(min_row=2):
            max_lines = 1
            for cell in row:
                column_letter = cell.column_letter
                
                # 根据列类型设置换行对齐
                if column_letter in numeric_columns:
                    cell.alignment = Alignment(wrap_text=True, horizontal="center", vertical="center")
                else:
                    cell.alignment = Alignment(wrap_text=True, horizontal="left", vertical="center")

                # 计算行高
                if cell.value and "\n" in str(cell.value):
                    lines_count = len(str(cell.value).split("\n"))
                    max_lines = max(max_lines, lines_count)

            # 设置行高
            if max_lines > 1:
                worksheet.row_dimensions[row[0].row].height = (
                    max_lines * self.config["row_height_factor"]
                )


    def _apply_excel_formatting(self, worksheet):
        """
        应用Excel格式化（修改版本，支持智能对齐）
        
        Args:
            worksheet: openpyxl工作表对象
        """
        # 检查工作表是否有数据
        if worksheet.max_row <= 1:
            return

        # 检测数字列
        numeric_columns = self._detect_numeric_columns(worksheet)
        print(f"🔢 检测到数字列: {numeric_columns} 在工作表: {worksheet.title}")

        # 设置标题行格式
        header_font = Font(name="Microsoft YaHei", bold=True, color=self.config["header_font_color"])
        header_fill = PatternFill(
            start_color=self.config["header_background_color"], 
            end_color=self.config["header_background_color"], 
            fill_type="solid"
        )
        
        # 格式化标题行
        for cell in worksheet[1]:
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center", vertical="center")  # 标题行居中
            cell.border = Border(
                left=Side(style="thin"),
                right=Side(style="thin"),
                top=Side(style="thin"),
                bottom=Side(style="thin"),
            )

        # 格式化数据行，根据列类型设置水平对齐
        for row in worksheet.iter_rows(min_row=2):
            for cell in row:
                column_letter = cell.column_letter
                
                # 设置字体
                cell.font = Font(name="Microsoft YaHei", bold=False)
                
                # 根据列类型设置对齐方式
                if column_letter in numeric_columns:
                    # 数字列：水平居中，垂直居中
                    cell.alignment = Alignment(horizontal="center", vertical="center")
                else:
                    # 文本列：左对齐，垂直居中
                    cell.alignment = Alignment(horizontal="left", vertical="center")
                
                # 设置边框
                cell.border = Border(
                    left=Side(style="thin", color="000000"),
                    right=Side(style="thin", color="000000"),
                    top=Side(style="thin", color="000000"),
                    bottom=Side(style="thin", color="000000"),
                )

        # 自动调整列宽和行高
        if self.config["auto_width"]:
            self._auto_adjust_columns(worksheet)

        if self.config["wrap_text"]:
            self._apply_text_wrapping(worksheet, numeric_columns)
            


    def convert_to_excel(
        self, data: List[Dict[str, Any]], output_path: str, verbose: bool = True
    ) -> Dict[str, Any]:
        """
        转换数据为Excel文件

        Args:
            data: 嵌套JSON数据列表
            output_path: 输出文件路径
            verbose: 是否显示详细信息

        Returns:
            转换结果统计信息
        """
        try:
            if verbose:
                print(f"🔄 开始转换 {len(data)} 条记录...")

            # 处理数据
            flattened_data = self.process_data(data, verbose)

            if not flattened_data:
                raise ValueError("处理后的数据为空")

            if verbose:
                print("📋 创建DataFrame...")

            # 创建DataFrame
            df_flat = pd.DataFrame(flattened_data)
            if verbose:
                print(f"✅ 对比数据表: {len(df_flat)} 行 × {len(df_flat.columns)} 列")

            # 创建原始数据DataFrame
            df_original = pd.DataFrame(
                [
                    {
                        self.config["index_column_name"]: i + 1,
                        self.config["json_column_name"]: json.dumps(
                            item, ensure_ascii=False, indent=2
                        ),
                    }
                    for i, item in enumerate(data)
                ]
            )
            if verbose:
                print(f"✅ 原始数据表: {len(df_original)} 行 × {len(df_original.columns)} 列")
                print(f"💾 写入Excel文件: {output_path}")

            # 写入Excel文件
            with pd.ExcelWriter(output_path, engine="openpyxl") as writer:
                # 写入对比数据
                if verbose:
                    print("📝 写入对比数据表...")
                df_flat.to_excel(
                    writer,
                    sheet_name=self.config["processed_sheet_name"],
                    index=False,
                )

                # 根据配置决定是否写入原始数据
                if self.config.get("output_original_data", True):
                    if verbose:
                        print("📝 写入原始数据表...")
                    df_original.to_excel(
                        writer,
                        sheet_name=self.config["original_sheet_name"],
                        index=False,
                    )

                # 应用格式化
                if verbose:
                    print("🎨 应用格式化...")

                # 格式化对比数据表
                flattened_ws = writer.sheets[self.config["processed_sheet_name"]]
                self._apply_excel_formatting(flattened_ws)

                # 格式化原始数据表（如果存在）
                if self.config.get("output_original_data", True):
                    original_ws = writer.sheets[self.config["original_sheet_name"]]
                    self._apply_excel_formatting(original_ws)

            if verbose:
                print("✅ Excel文件创建成功!")

            # 返回统计信息
            result = {
                "success": True,
                "output_path": output_path,
                "records_count": len(data),
                "flattened_columns": len(df_flat.columns),
                "column_names": list(df_flat.columns),
                "sheets": [
                    self.config["processed_sheet_name"],
                ],
            }

            # 如果输出原始数据表，则添加其名称
            if self.config.get("output_original_data", True):
                result["sheets"].append(self.config["original_sheet_name"])

            return result

        except Exception as e:
            if verbose:
                print(f"❌ 转换失败: {str(e)}")
                traceback.print_exc()
            return {"success": False, "error": str(e), "output_path": output_path}

    # 简化的接口方法
    def convert(self, data, output_path, **kwargs):
        """
        简化的转换方法
        
        Args:
            data: JSON数据（列表或单个字典）
            output_path: 输出文件路径
            **kwargs: 其他参数
        
        Returns:
            转换结果
        """
        if isinstance(data, dict):
            data = [data]
        
        return self.convert_to_excel(data, output_path, **kwargs)

    def merge_excel_files(self, excel_files, output_path):
        """
        合并多个Excel文件
        """
        # 创建一个Excel写入器
        with pd.ExcelWriter(output_path, engine="openpyxl") as writer:
            sheet_names_count = {}
            for file in excel_files:
                xls = pd.ExcelFile(file)
                for sheet_name in xls.sheet_names:
                    df = pd.read_excel(file, sheet_name=sheet_name)
                    
                    # 处理重复的sheet名称
                    if sheet_name in sheet_names_count:
                        sheet_names_count[sheet_name] += 1
                        new_sheet_name = f"{sheet_name}_{sheet_names_count[sheet_name]}"
                    else:
                        sheet_names_count[sheet_name] = 1
                        new_sheet_name = sheet_name
                    
                    df.to_excel(writer, sheet_name=new_sheet_name, index=False)  # 将每个sheet写入同一个Excel文件

        # 读取合并后的Excel文件
        with pd.ExcelWriter(output_path, engine="openpyxl", mode='a') as writer:
            # 重新读取合并后的Excel文件以获取工作表
            merged_workbook = writer.book
            for sheet_name in merged_workbook.sheetnames:
                worksheet = merged_workbook[sheet_name]
                self._apply_excel_formatting(worksheet)  # 应用格式化

        return output_path