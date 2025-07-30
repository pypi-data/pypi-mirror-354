"""Excel utils.

Developing based on `openpyxl` and `xlsxwriter`.
"""

from __future__ import annotations

import logging
from collections import defaultdict
from colorsys import hls_to_rgb, rgb_to_hls
from dataclasses import dataclass
from pathlib import Path
from typing import Any, Iterable, Literal

import polars as pl
from typing_extensions import Self

try:
    import openpyxl
    import xlsxwriter
    import xlsxwriter.format
    import xlsxwriter.worksheet
    from openpyxl.cell.cell import Cell, MergedCell
    from openpyxl.styles.colors import Color
    from openpyxl.utils import coordinate_to_tuple, get_column_letter
    from openpyxl.worksheet.worksheet import Worksheet
    from openpyxl.xml.functions import QName, fromstring
    from xlsxwriter.utility import xl_cell_to_rowcol, xl_rowcol_to_cell
except ModuleNotFoundError as _:
    raise ModuleNotFoundError(
        "The 'excel' module requires optional dependencies 'openpyxl' and 'xlsxwriter'. "
        "You can install them using the following command:\n"
        'pip install "kaitian[excel]"'
    )


@dataclass
class TextFormat:
    font_name: str | None
    font_size: float = 11
    font_color: str = "#000000"
    bold: bool = False
    italic: bool = False
    underline: Literal["single", "double", "single_full", "double_full"] | None = None
    align: str | None = None
    valign: str | None = None
    background_color: str | None = None
    text_wrap: bool = False
    num_format: str = "General"

    @staticmethod
    def _underline_to_writer(underline: str | None) -> int | None:
        if underline is None:
            return None

        if underline == "double":
            return 2
        elif underline == "single_full":
            return 33
        elif underline == "double_full":
            return 34
        return 1

    def to_writer(self) -> dict:
        fmt_all = {
            "font_name": self.font_name,
            "font_size": self.font_size,
            "font_color": self.font_color,
            "bold": self.bold,
            "italic": self.italic,
            "underline": self._underline_to_writer(self.underline),
            "align": self.align,
            "valign": self.valign,
            "bg_color": self.background_color,
            "text_wrap": self.text_wrap,
            "num_format": self.num_format,
        }
        return {k: v for k, v in fmt_all.items() if v is not None}


@dataclass
class BorderFormat:
    direction: str | Literal["top", "bottom", "left", "right", "all"]
    style: str = "thin"
    color: str | None = None

    @staticmethod
    def _style2idx(style: str) -> int:
        if style == "thin":
            return 1
        elif style == "double":
            return 5
        elif style == "medium":
            return 2
        elif style == "mediumDashed":
            return 8
        else:
            return 1

    def to_writer(self) -> dict:
        if self.direction == "all":
            return {"border": self._style2idx(self.style), "border_color": self.color}
        else:
            return {
                f"{self.direction}": self._style2idx(self.style),
                f"{self.direction}_color": self.color,
            }


@dataclass
class SheetFormat:
    """Sheet-level format"""

    cell_range: str | tuple[int, int, int, int]  # A1:B2 or (1, 1, 2, 2)
    ftype: Literal["data_bar", "color_scale"]
    fopt: dict | None = None

    def to_writer(self) -> dict:
        if isinstance(self.cell_range, str):
            cell_range = (
                *(xl_cell_to_rowcol(self.cell_range.split(":")[0])),
                *(xl_cell_to_rowcol(self.cell_range.split(":")[1])),
            )
        else:
            cell_range = self.cell_range

        fmt: dict[str, Any] = {
            "first_row": cell_range[0],
            "first_col": cell_range[1],
            "last_row": cell_range[2],
            "last_col": cell_range[3],
        }

        if self.ftype == "data_bar":
            fmt["options"] = {"type": "data_bar", **(self.fopt or {})}
        elif self.ftype == "color_scale":
            opt_name = list(
                map(lambda x: x.split("_")[0], list((self.fopt or {}).keys()))
            )
            if "mid" in opt_name:
                fmt["options"] = {"type": "3_color_scale", **(self.fopt or {})}
            else:
                fmt["options"] = {"type": "2_color_scale", **(self.fopt or {})}

        return fmt


class ThemeParser:
    HLSMAX = 240
    XLMNS = "http://schemas.openxmlformats.org/drawingml/2006/main"

    def __init__(self, workbook: openpyxl.Workbook | Path | str) -> None:
        if isinstance(workbook, (Path, str)):
            self.workbook = openpyxl.load_workbook(Path(workbook), read_only=True)
        else:
            self.workbook = workbook

        self._parse_theme_color()

    def _parse_theme_color(self) -> None:
        _color_scheme = (
            fromstring(self.workbook.loaded_theme)
            .find(QName(self.XLMNS, "themeElements").text)
            .findall(QName(self.XLMNS, "clrScheme").text)[0]
        )

        self.colors = []

        for _c in [
            "lt1",
            "dk1",
            "lt2",
            "dk2",
            "accent1",
            "accent2",
            "accent3",
            "accent4",
            "accent5",
            "accent6",
        ]:
            accent = _color_scheme.find(QName(self.XLMNS, _c).text)
            for i in list(accent):  # walk all child nodes, rather than assuming [0]
                if "window" in i.attrib["val"]:
                    self.colors.append(i.attrib["lastClr"])
                else:
                    self.colors.append(i.attrib["val"])

    def _rgb2hls(self, rgb: str) -> tuple[int, int, int]:
        """From RGB to HLS
        Converts rgb values in range (0,1) or a hex string of the form
        '[#aa]rrggbb' to HLSMAX based HLS, alpha values are ignored.
        """
        if len(rgb) > 6:
            rgb = rgb[-6:]
        red = int(rgb[0:2], 16) / 255
        green = int(rgb[2:4], 16) / 255
        blue = int(rgb[4:6], 16) / 255

        _h, _l, _s = rgb_to_hls(red, green, blue)
        return (
            int(round(_h * self.HLSMAX)),
            int(round(_l * self.HLSMAX)),
            int(round(_s * self.HLSMAX)),
        )

    def get_theme_color(self, theme: int = 0, tint: float = 0.0) -> str:
        argb_main = self.colors[theme]
        hue, lightness, saturation = self._rgb2hls(argb_main)

        # See: http://ciintelligence.blogspot.co.uk/2012/02/converting-excel-theme-color-and-tint.html
        if tint < 0:
            lightness = int(round(lightness * (1.0 + tint)))
        else:
            lightness = int(
                round(
                    lightness * (1.0 - tint)
                    + (self.HLSMAX - self.HLSMAX * (1.0 - tint))
                )
            )

        r, g, b = hls_to_rgb(
            hue / self.HLSMAX, lightness / self.HLSMAX, saturation / self.HLSMAX
        )
        return (
            "#%02x%02x%02x"
            % (
                int(round(r * 255)),
                int(round(g * 255)),
                int(round(b * 255)),
            )
        ).upper()

    def get_indexed_color(self, index: int = 0) -> str:
        color_idx_map = openpyxl.styles.colors.COLOR_INDEX  # type: ignore
        return f"#{color_idx_map[index % len(color_idx_map)][2:]}"


class ExcelReader:
    def __init__(self, workbook: openpyxl.Workbook | Path | str) -> None:
        if isinstance(workbook, (Path, str)):
            self.workbook = openpyxl.load_workbook(Path(workbook))
        else:
            self.workbook = workbook

        # theme parser
        self.workbook_theme = ThemeParser(self.workbook)
        self._logger = logging.getLogger(__name__)

    def _color2rgb(self, color: Color | None, default: str | None = None) -> str | None:
        if default is None:
            default_rgb = None
        elif default.lower() == "white":
            default_rgb = "#FFFFFF"
        elif default.lower() == "black":
            default_rgb = "#000000"
        else:
            default_rgb = default

        if color is None:
            return default_rgb

        if color.type == "rgb":
            if color.rgb == "00000000":
                return default_rgb
            return f"#{color.rgb[2:]}"
        elif color.type == "theme":
            return self.workbook_theme.get_theme_color(color.theme, color.tint)
        elif color.type == "indexed":
            return self.workbook_theme.get_indexed_color(color.indexed)
        else:
            self._logger.warning(f"Unknown color type: {color.type}")
            return default_rgb

    def _get_worksheet(self, sheet: str | int) -> Worksheet:
        if isinstance(sheet, str):
            worksheet = self.workbook[sheet]
        elif isinstance(sheet, int):
            worksheet = self.workbook.worksheets[sheet]
        else:
            raise ValueError("sheet must be str or int")

        return worksheet

    def _get_position(self, position: str | tuple[int, int]) -> tuple[int, int]:
        # 1-based indexing
        if isinstance(position, str):
            return coordinate_to_tuple(position)
        else:
            return position

    def _get_textfmt(self, workcell: Cell | MergedCell) -> TextFormat:
        font = workcell.font
        text_fmt = {
            "font_name": font.name,
            "font_size": font.size,
            "font_color": self._color2rgb(font.color, default="black"),
            "bold": font.bold,
            "italic": font.italic,
            "underline": font.underline,
            "align": workcell.alignment.horizontal,
            "valign": workcell.alignment.vertical,
            "text_wrap": workcell.alignment.wrap_text or False,
            "num_format": workcell.number_format,
            "background_color": self._color2rgb(workcell.fill.fgColor),
        }

        return TextFormat(**text_fmt)

    def _get_shape(
        self, worksheet: Worksheet, position: tuple[int, int]
    ) -> tuple[float, float]:
        coldim = worksheet.column_dimensions[get_column_letter(position[0])]
        rowdim = worksheet.row_dimensions[position[1]]
        if getattr(coldim, "width") is not None:
            width = coldim.width
        else:
            width = coldim.style

        if getattr(rowdim, "height") is not None:
            height = rowdim.height
        else:
            height = rowdim.s

        return width, height

    def _get_border(self, workcell: Cell | MergedCell) -> list[BorderFormat]:
        # cell border of (style, color)
        borders = []
        for direction in ["left", "right", "top", "bottom"]:
            if getattr(workcell.border, direction).style is not None:
                borders.append(
                    BorderFormat(
                        direction,
                        getattr(workcell.border, direction).style,
                        self._color2rgb(getattr(workcell.border, direction).color),
                    )
                )
        return borders

    def _get_conditional(self, workcell: Cell | MergedCell) -> dict[str, Any]:
        # TODO: 解析条件格式
        raise NotImplementedError("get_conditional formatter not implemented yet.")

    def get_cell(self, sheet: str | int, cell: str | tuple[int, int]) -> dict[str, Any]:
        worksheet = self._get_worksheet(sheet)
        position = self._get_position(cell)
        max_row, max_col = worksheet.max_row, worksheet.max_column
        if position[0] > max_row or position[1] > max_col:
            self._logger.warning(
                f"The cell {cell} is out of range, max_row={max_row}, max_col={max_col}"
            )
            return {}

        workcell = worksheet.cell(*position)
        value = workcell.value
        return {
            "value": value,
            "text_format": self._get_textfmt(workcell),
            "border_format": self._get_border(workcell),
            "shape": self._get_shape(worksheet, position),
        }


class ExcelWriter:
    def __init__(self, file_path: str | Path) -> None:
        self.workbook = xlsxwriter.Workbook(Path(file_path))
        self.formatters: dict[str, xlsxwriter.format.Format] = {}
        self.default_textfmt: TextFormat | None = None
        self.default_borderfmt: BorderFormat | None = None
        self.sheet_offset: defaultdict[str, str] = defaultdict(lambda: "A1")

    @property
    def default_formatter(self) -> list[TextFormat | BorderFormat]:
        default_fmt = []
        if self.default_textfmt is not None:
            default_fmt.append(self.default_textfmt)
        if self.default_borderfmt is not None:
            default_fmt.append(self.default_borderfmt)
        return default_fmt

    def _get_worksheet_or_create(
        self, sheet_name: str
    ) -> xlsxwriter.worksheet.Worksheet:
        if sheet_name in self.workbook.sheetnames:
            worksheet = self.workbook.get_worksheet_by_name(sheet_name)
        else:
            worksheet = self.workbook.add_worksheet(sheet_name)

        assert worksheet is not None
        return worksheet

    def _parse_format(
        self,
        text_fmt: TextFormat | list[TextFormat] | None = None,
        border_fmt: BorderFormat | list[BorderFormat] | None = None,
    ) -> xlsxwriter.format.Format:
        if isinstance(text_fmt, TextFormat):
            text_fmt = [text_fmt]

        text_fmt_writer = {}
        if text_fmt is not None:
            for tfmt in text_fmt:
                text_fmt_writer.update(tfmt.to_writer())

        if isinstance(border_fmt, BorderFormat):
            border_fmt = [border_fmt]

        border_fmt_writer = {}
        if border_fmt is not None:
            for bfmt in border_fmt:
                text_fmt_writer.update(bfmt.to_writer())

        return self.workbook.add_format({**text_fmt_writer, **border_fmt_writer})

    def add_format(
        self,
        name: str | None = None,
        text_fmt: TextFormat | None = None,
        border_fmt: BorderFormat | None = None,
    ) -> Self:
        if name is None:
            self.default_textfmt = text_fmt
            self.default_borderfmt = border_fmt
        else:
            self.formatters[name] = self._parse_format(text_fmt, border_fmt)
        return self

    def get_format(
        self,
        cell_format: str
        | TextFormat
        | BorderFormat
        | list[TextFormat | BorderFormat]
        | None,
    ) -> xlsxwriter.format.Format:
        fmt = None
        if isinstance(cell_format, str):
            fmt = self.formatters.get(cell_format)
        else:
            if isinstance(cell_format, (TextFormat, BorderFormat)):
                cell_format = [*self.default_formatter, cell_format]
            elif isinstance(cell_format, list):
                cell_format = [*self.default_formatter, *cell_format]
            else:
                cell_format = [*self.default_formatter]

            _cell_fmt_text = [fmt for fmt in cell_format if isinstance(fmt, TextFormat)]
            _cell_fmt_border = [
                fmt for fmt in cell_format if isinstance(fmt, BorderFormat)
            ]
            fmt = self._parse_format(_cell_fmt_text or None, _cell_fmt_border or None)

        if fmt is None:
            return self.workbook.add_format({})
        else:
            return fmt

    def fill_cell(
        self,
        sheet_name: str,
        position: str | tuple[int, int],
        value: Any,
        cell_format: str
        | TextFormat
        | BorderFormat
        | list[TextFormat | BorderFormat]
        | None = None,
    ) -> Self:
        if isinstance(position, tuple):
            position = xl_rowcol_to_cell(*position)
        worksheet = self._get_worksheet_or_create(sheet_name)
        fmt = self.get_format(cell_format)
        worksheet.write(position, value, fmt)
        return self

    def fill_row(
        self,
        sheet_name: str,
        position: str | tuple[int, int],
        value: Iterable,
        cell_format: str
        | TextFormat
        | BorderFormat
        | list[TextFormat | BorderFormat]
        | None = None,
    ) -> Self:
        if isinstance(position, str):
            position = xl_cell_to_rowcol(position)
        worksheet = self._get_worksheet_or_create(sheet_name)
        fmt = self.get_format(cell_format)
        worksheet.write_row(*position, value, fmt)
        return self

    def fill_column(
        self,
        sheet_name: str,
        position: str | tuple[int, int],
        value: Iterable,
        cell_format: str
        | TextFormat
        | BorderFormat
        | list[TextFormat | BorderFormat]
        | None = None,
    ) -> Self:
        if isinstance(position, str):
            position = xl_cell_to_rowcol(position)
        worksheet = self._get_worksheet_or_create(sheet_name)
        fmt = self.get_format(cell_format)
        worksheet.write_column(*position, value, fmt)
        return self

    def fill_merge(
        self,
        sheet_name: str,
        offset: str | tuple[int, int],
        nrows: int,
        ncols: int,
        value: Any = None,
        cell_format: str
        | TextFormat
        | BorderFormat
        | list[TextFormat | BorderFormat]
        | None = None,
    ) -> Self:
        if isinstance(offset, str):
            offset = xl_cell_to_rowcol(offset)

        fmt = self.get_format(cell_format)
        worksheet = self._get_worksheet_or_create(sheet_name)
        _ = worksheet.merge_range(
            *offset,
            last_row=offset[0] + nrows - 1,
            last_col=offset[1] + ncols - 1,
            data=value,
            cell_format=fmt,
        )
        return self

    def set_height(self, sheet_name: str, row: int, height: float) -> Self:
        worksheet = self._get_worksheet_or_create(sheet_name)
        worksheet.set_row(row - 1, height)
        return self

    def set_width(self, sheet_name: str, col: str, width: float) -> Self:
        worksheet = self._get_worksheet_or_create(sheet_name)
        _, col_idx = xl_cell_to_rowcol(f"{col}1")
        worksheet.set_column(first_col=col_idx, last_col=col_idx, width=width)
        return self

    def set_sheet_format(
        self, sheet_name, sheet_format: SheetFormat | list[SheetFormat]
    ) -> Self:
        worksheet = self._get_worksheet_or_create(sheet_name)
        if isinstance(sheet_format, SheetFormat):
            sheet_format = [sheet_format]

        for fmt in sheet_format:
            worksheet.conditional_format(**fmt.to_writer())

        return self

    def write_table(
        self,
        sheet_name: str,
        data: pl.DataFrame,
        title: str | None = None,
        header: bool = True,
        index: bool = False,
        offset: str | tuple[int, int] | None = None,
        title_format: str
        | TextFormat
        | BorderFormat
        | list[TextFormat | BorderFormat]
        | None = None,
        header_format: str
        | TextFormat
        | BorderFormat
        | list[TextFormat | BorderFormat]
        | None = None,
        data_format: str
        | TextFormat
        | BorderFormat
        | list[TextFormat | BorderFormat]
        | dict[
            str,
            str | TextFormat | BorderFormat | list[TextFormat | BorderFormat] | None,
        ]
        | None = None,
    ) -> Self:
        if index:
            data = data.with_row_index(offset=1, name="序号")

        if not isinstance(data_format, dict):
            data_format = {k: data_format for k in data.columns}

        # worksheet = self._get_worksheet_or_create(sheet_name)

        if offset is None:
            offset_rc = xl_cell_to_rowcol(self.sheet_offset[sheet_name])
        elif isinstance(offset, str):
            offset_rc: tuple[int, int] = xl_cell_to_rowcol(offset)
        else:
            offset_rc = offset

        if title is not None:
            _ = self.fill_merge(
                sheet_name,
                offset_rc,
                nrows=1,
                ncols=data.width,
                value=title,
                cell_format=title_format,
            )
            offset_rc = (offset_rc[0] + 1, offset_rc[1])

        if header:
            _ = self.fill_row(
                sheet_name, offset_rc, data.columns, cell_format=header_format
            )
            offset_rc = (offset_rc[0] + 1, offset_rc[1])

        for col_idx, series in enumerate(data):
            _ = self.fill_column(
                sheet_name,
                (offset_rc[0], offset_rc[1] + col_idx),
                series.to_list(),
                cell_format=data_format.get(series.name),
            )
        offset_rc = (offset_rc[0] + data.height + 1, offset_rc[1])
        self.sheet_offset[sheet_name] = xl_rowcol_to_cell(*offset_rc)
        return self

    def done(self) -> None:
        self.workbook.close()
